"""
Multi-document RAG evaluation script.

Loads 3 documents into a single Weaviate collection and proves that ancestral-
heading-aware retrieval outperforms text-only retrieval using Precision@K.

Key design:
  - doc_source is derived from the ROOT heading in ancestral_headings (not filename)
  - "With Ancestral"   : combined score = w_text * text_score + w_heading * heading_score
  - "Without Ancestral": score = text_score only (heading vector ignored)
  - Precision@K = (# top-K results from the target document) / K

Usage (via main.py):
    python database/main.py --test multidoc

Usage (standalone):
    python database/tests/eval_multidoc.py
"""

import time
import warnings
import math
import unicodedata
import json
import sys
import io
from pathlib import Path
from collections import Counter

import requests
import weaviate
from weaviate.classes.config import Configure, Property, DataType
from weaviate.classes.query import MetadataQuery, HybridFusion, TargetVectors
from weaviate.config import AdditionalConfig, Timeout
from weaviate.exceptions import WeaviateQueryError
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

# ========================
# PURE HELPERS (no config dependency)
# ========================
def _clean(s: str) -> str:
    if not s:
        return ""
    s = unicodedata.normalize("NFKC", s)
    s = "".join(c for c in s if unicodedata.category(c) not in ("Cc", "Cf") or c in "\t\n\r")
    return s.strip()


def build_heading_path(chunk_data: dict) -> str:
    """Build the richest heading path from ancestral_headings for the heading vector."""
    anc = chunk_data.get("ancestral_headings", {})
    if not isinstance(anc, dict):
        return ""
    paths = []
    for heading, ancestors in anc.items():
        if isinstance(ancestors, list) and ancestors:
            paths.append(" > ".join(ancestors + [heading]))
        else:
            paths.append(heading)
    if paths:
        paths.sort(key=lambda p: p.count(" > "), reverse=True)
        return paths[0]
    return ""


def get_doc_source(chunk_data: dict) -> str:
    """
    Return the root document name from ancestral_headings.
    The root heading is the one whose ancestors list is empty (depth = 0).
    Falls back to the first ancestor if no root is found.
    """
    anc = chunk_data.get("ancestral_headings", {})
    if not isinstance(anc, dict) or not anc:
        return "UNKNOWN"
    for heading, ancestors in anc.items():
        if isinstance(ancestors, list) and len(ancestors) == 0:
            return heading
    for heading, ancestors in anc.items():
        if isinstance(ancestors, list) and ancestors:
            return ancestors[0]
    return "UNKNOWN"


# ========================
# MAIN ENTRY POINT
# ========================
def run(cfg: dict) -> None:
    """
    Execute the multi-document evaluation.

    Args:
        cfg: Merged config dict — contains top-level weaviate/ollama keys
             plus eval_multidoc section keys (retrieval, eval_data_file, etc.).
             main.py passes {**full_cfg, **full_cfg["eval_multidoc"]}.
    """
    TESTS_DIR = Path(__file__).resolve().parent
    BASE_DIR  = TESTS_DIR.parent  # database/ — for output/ paths

    warnings.filterwarnings("ignore", category=DeprecationWarning, module="weaviate")

    OLLAMA_HOST      = cfg["ollama"]["host"]
    OLLAMA_DOCKER_EP = cfg["ollama"]["docker_endpoint"]
    EMBED_MODEL      = cfg["ollama"]["embed_model"]
    GENERATIVE_MODEL = cfg["ollama"]["generative_model"]
    W_TEXT           = cfg["retrieval"]["w_text"]
    W_HEADING        = cfg["retrieval"]["w_heading"]
    TOP_K_RETRIEVE   = cfg["retrieval"]["top_k_retrieve"]
    TOP_N            = cfg["retrieval"]["top_n"]
    PRECISION_K      = cfg["retrieval"].get("precision_k", TOP_N)
    DATA_FILE        = cfg["eval_data_file"]
    EXCEL_OUTPUT_FILE  = cfg.get("excel_output_file", "output/multidoc_eval_results.xlsx")
    QUERIES_FILE       = cfg.get("queries_file", "eval_queries.json")

    data_path  = Path(DATA_FILE)
    excel_path = BASE_DIR / EXCEL_OUTPUT_FILE

    def _embed_direct(text: str):
        try:
            r = requests.post(
                f"{OLLAMA_HOST}/api/embed",
                json={"model": EMBED_MODEL, "input": text or "document"},
                timeout=60,
            )
            r.raise_for_status()
            vec = r.json().get("embeddings", [[]])[0]
            if any(math.isnan(v) or math.isinf(v) for v in vec):
                return None
            return vec
        except Exception:
            return None

    # ========================
    # TEST QUERIES — loaded from eval_queries.json (in tests/)
    # ========================
    _queries_path = TESTS_DIR / QUERIES_FILE
    if not _queries_path.exists():
        raise FileNotFoundError(f"Queries file not found: {_queries_path}\nCheck 'queries_file' in config.json.")

    with open(_queries_path, "r", encoding="utf-8") as _qf:
        TEST_QUERIES = json.load(_qf)

    print(f"Loaded {len(TEST_QUERIES)} test queries from: {_queries_path}")

    # ========================
    # PRE-FLIGHT CHECKS
    # ========================
    if not data_path.exists():
        raise FileNotFoundError(f"Data file not found: {data_path}\nCheck 'eval_data_file' in config.json.")

    if excel_path.exists():
        raise FileExistsError(
            f"Excel output already exists: {excel_path}\n"
            "Rename or delete it before re-running."
        )

    print(f"Loading data from: {data_path}")
    with open(data_path, "r", encoding="utf-8") as _f:
        data = json.load(_f)

    chunk_doc_map: dict[str, str] = {}
    for item in data:
        for k, v in item.items():
            if k.startswith("chunk_id") and isinstance(v, dict):
                chunk_doc_map[k] = get_doc_source(v)

    print(f"Loaded {len(chunk_doc_map)} chunks from {len(set(chunk_doc_map.values()))} documents:")
    for doc, cnt in sorted(Counter(chunk_doc_map.values()).items()):
        print(f"  [{cnt:2d} chunks]  {doc}")

    # ========================
    # CONNECT TO WEAVIATE & INSERT
    # ========================
    print("\nConnecting to Weaviate...")
    client = weaviate.connect_to_custom(
        http_host=cfg["weaviate"]["http_host"],
        http_port=cfg["weaviate"]["http_port"],
        http_secure=False,
        grpc_host=cfg["weaviate"]["grpc_host"],
        grpc_port=cfg["weaviate"]["grpc_port"],
        grpc_secure=False,
        skip_init_checks=True,
        additional_config=AdditionalConfig(timeout=Timeout(init=30, query=120, insert=120)),
    )

    try:
        if client.collections.exists("LangExtractEval"):
            client.collections.delete("LangExtractEval")

        client.collections.create(
            name="LangExtractEval",
            vector_config=[
                Configure.Vectors.text2vec_ollama(
                    name="text_vector",
                    source_properties=["text"],
                    api_endpoint=OLLAMA_DOCKER_EP,
                    model=EMBED_MODEL,
                ),
                Configure.Vectors.text2vec_ollama(
                    name="heading_vector",
                    source_properties=["heading_path"],
                    api_endpoint=OLLAMA_DOCKER_EP,
                    model=EMBED_MODEL,
                ),
            ],
            generative_config=Configure.Generative.ollama(
                api_endpoint=OLLAMA_DOCKER_EP,
                model=GENERATIVE_MODEL,
            ),
            properties=[
                Property(name="text",              data_type=DataType.TEXT),
                Property(name="heading_path",      data_type=DataType.TEXT),
                Property(name="chunk_id",          data_type=DataType.TEXT, skip_vectorization=True),
                Property(name="doc_source",        data_type=DataType.TEXT, skip_vectorization=True),
                Property(name="ancestral_headings", data_type=DataType.TEXT, skip_vectorization=True),
            ],
        )
        print("Collection 'LangExtractEval' created.")

        collection = client.collections.use("LangExtractEval")

        inserted_total = 0
        for item in data:
            for chunk_id, chunk_data in item.items():
                if not (chunk_id.startswith("chunk_id") and isinstance(chunk_data, dict)):
                    continue

                raw_text   = item.get("Text", "")
                text       = _clean(str(raw_text) if raw_text is not None else "")
                heading_path = _clean(build_heading_path(chunk_data))
                doc_source   = get_doc_source(chunk_data)
                anc_str      = json.dumps(chunk_data.get("ancestral_headings", {}))

                if not heading_path:
                    heading_path = doc_source or f"Section: {chunk_id}"

                obj = {
                    "text":               text,
                    "heading_path":       heading_path,
                    "chunk_id":           chunk_id,
                    "doc_source":         doc_source,
                    "ancestral_headings": anc_str,
                }

                inserted = False
                for attempt in range(3):
                    try:
                        collection.data.insert(obj)
                        inserted = True
                        break
                    except Exception:
                        if attempt == 0:
                            time.sleep(1)
                            obj = {**obj, "text": obj["text"][:512], "heading_path": obj["heading_path"][:100]}
                        elif attempt == 1:
                            time.sleep(2)
                            safe_t = "".join(c for c in obj["text"][:256] if ord(c) < 128).strip() or "document content"
                            safe_h = "".join(c for c in obj["heading_path"] if ord(c) < 128).strip() or "document section"
                            obj = {**obj, "text": safe_t, "heading_path": safe_h}
                        else:
                            t_vec = _embed_direct(obj["text"]) or _embed_direct("document content")
                            h_vec = _embed_direct(obj["heading_path"]) or _embed_direct("document section")
                            if t_vec and h_vec:
                                try:
                                    collection.data.insert(obj, vector={"text_vector": t_vec, "heading_vector": h_vec})
                                    inserted = True
                                except Exception:
                                    pass

                if inserted:
                    inserted_total += 1
                else:
                    print(f"  [WARN] Could not insert {chunk_id}")

        print(f"Inserted {inserted_total}/{len(chunk_doc_map)} chunks. Waiting for vectorization...")
        time.sleep(10)

        # ========================
        # RETRIEVAL FUNCTION
        # ========================
        def run_retrieval(query_text: str, include_ancestral: bool) -> list[dict]:
            search_text = None
            for attempt in range(3):
                try:
                    search_text = collection.query.hybrid(
                        query=query_text,
                        target_vector=TargetVectors.sum(["text_vector"]),
                        query_properties=["text"],
                        alpha=0.7,
                        fusion_type=HybridFusion.RELATIVE_SCORE,
                        limit=TOP_K_RETRIEVE,
                        return_metadata=MetadataQuery(score=True),
                    )
                    break
                except WeaviateQueryError:
                    if attempt < 2:
                        time.sleep(2 * (attempt + 1))

            search_heading = None
            if include_ancestral:
                for attempt in range(3):
                    try:
                        search_heading = collection.query.hybrid(
                            query=query_text,
                            target_vector=TargetVectors.sum(["heading_vector"]),
                            query_properties=["heading_path"],
                            alpha=0.7,
                            fusion_type=HybridFusion.RELATIVE_SCORE,
                            limit=TOP_K_RETRIEVE,
                            return_metadata=MetadataQuery(score=True),
                        )
                        break
                    except WeaviateQueryError:
                        if attempt < 2:
                            time.sleep(2 * (attempt + 1))

            if not search_text or not getattr(search_text, "objects", None):
                return []

            text_scores:    dict[str, float] = {}
            heading_scores: dict[str, float] = {}
            obj_map:        dict[str, object] = {}

            for o in search_text.objects:
                cid = o.properties.get("chunk_id", "")
                s   = o.metadata.score if (o.metadata and o.metadata.score is not None) else 0.0
                text_scores[cid] = max(text_scores.get(cid, 0.0), s)
                obj_map[cid] = o

            if search_heading and getattr(search_heading, "objects", None):
                for o in search_heading.objects:
                    cid = o.properties.get("chunk_id", "")
                    s   = o.metadata.score if (o.metadata and o.metadata.score is not None) else 0.0
                    heading_scores[cid] = max(heading_scores.get(cid, 0.0), s)
                    if cid not in obj_map:
                        obj_map[cid] = o

            all_cids = set(list(text_scores.keys()) + list(heading_scores.keys()))
            max_t = max(text_scores.values(), default=1.0) or 1.0
            max_h = max(heading_scores.values(), default=1.0) or 1.0

            rows = []
            for cid in all_cids:
                t        = text_scores.get(cid, 0.0) / max_t
                h        = (heading_scores.get(cid, 0.0) / max_h) if include_ancestral else 0.0
                combined = (W_TEXT * t + W_HEADING * h) if include_ancestral else t
                obj      = obj_map.get(cid)
                doc      = (obj.properties.get("doc_source", "") if obj else "") or chunk_doc_map.get(cid, "UNKNOWN")
                rows.append({
                    "chunk_id":           cid,
                    "doc_source":         doc,
                    "text_score_norm":    round(t, 6),
                    "heading_score_norm": round(h, 6),
                    "combined_score":     round(combined, 6),
                })

            rows.sort(key=lambda r: r["combined_score"], reverse=True)
            return [{"rank": rank, **row} for rank, row in enumerate(rows[:TOP_N], 1)]

        # ========================
        # EVALUATE ALL QUERIES
        # ========================
        all_results = []
        k = min(PRECISION_K, TOP_N)
        total = len(TEST_QUERIES)

        print(f"\nRunning {total} queries (Precision@{k}) ...")
        print("-" * 70)

        for idx, tq in enumerate(TEST_QUERIES, 1):
            print(f"[{idx}/{total}] {tq['query_id']} | {tq['question'][:65]}...")

            with_anc    = run_retrieval(tq["question"], include_ancestral=True)
            without_anc = run_retrieval(tq["question"], include_ancestral=False)

            p_with    = sum(1 for r in with_anc[:k]    if r["doc_source"] == tq["target_doc"]) / k
            p_without = sum(1 for r in without_anc[:k] if r["doc_source"] == tq["target_doc"]) / k
            delta     = round(p_with - p_without, 4)

            all_results.append({
                **tq,
                "with_ancestral":    with_anc,
                "without_ancestral": without_anc,
                "precision_with":    round(p_with, 4),
                "precision_without": round(p_without, 4),
                "precision_delta":   delta,
            })

            marker = "✓ ANCESTRAL WINS" if delta > 0 else ("= TIE" if delta == 0 else "✗ TEXT WINS")
            print(f"  P@{k}: With={p_with:.0%}  Without={p_without:.0%}  Delta={delta:+.0%}  {marker}")

        avg_with    = sum(r["precision_with"]    for r in all_results) / len(all_results)
        avg_without = sum(r["precision_without"] for r in all_results) / len(all_results)
        avg_delta   = avg_with - avg_without

        print("-" * 70)
        print(f"AVERAGE Precision@{k}:")
        print(f"  With Ancestral   : {avg_with:.2%}")
        print(f"  Without Ancestral: {avg_without:.2%}")
        print(f"  Delta            : {avg_delta:+.2%}")
        print("-" * 70)

        # ========================
        # WRITE EXCEL
        # ========================
        wb = Workbook()

        hdr_font     = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill_blu = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hdr_fill_grn = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
        hdr_fill_orn = PatternFill(start_color="833C00", end_color="833C00", fill_type="solid")
        sub_font     = Font(bold=True, size=10)
        sub_fill     = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        thin_border  = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        gold_fill  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        ctr        = Alignment(horizontal="center", vertical="center")
        wrap_top   = Alignment(wrap_text=True, vertical="top")

        ws1 = wb.active
        ws1.title = "Detailed Results"

        ws1.merge_cells("A1:C1");  ws1["A1"] = "Query Info"
        ws1.merge_cells("D1:J1");  ws1["D1"] = f"With Ancestral-Heading  (w_text={W_TEXT}, w_heading={W_HEADING})"
        ws1.merge_cells("K1:O1");  ws1["K1"] = "Without Ancestral-Heading  (text score only)"
        for ref, fill in [("A1", hdr_fill_blu), ("D1", hdr_fill_grn), ("K1", hdr_fill_orn)]:
            c = ws1[ref]; c.font = hdr_font; c.fill = fill; c.alignment = ctr; c.border = thin_border

        col_hdrs = [
            "Q#", "Question", "Target Doc",
            "rank", "chunk_id", "doc_source", "Correct?", "text_norm", "heading_norm", "combined",
            "rank", "chunk_id", "doc_source", "Correct?", "text_norm",
        ]
        for ci, h in enumerate(col_hdrs, 1):
            c = ws1.cell(row=2, column=ci, value=h)
            c.font = sub_font; c.fill = sub_fill; c.border = thin_border; c.alignment = ctr

        cur_row = 3
        for res in all_results:
            n_rows = max(len(res["with_ancestral"]), len(res["without_ancestral"]), 1)

            ws1.cell(row=cur_row, column=1, value=res["query_id"]).alignment = ctr
            ws1.cell(row=cur_row, column=2, value=res["question"]).alignment = wrap_top
            ws1.cell(row=cur_row, column=3, value=res["target_doc"]).alignment = wrap_top

            for i in range(n_rows):
                row = cur_row + i

                if i < len(res["with_ancestral"]):
                    wa      = res["with_ancestral"][i]
                    correct = wa["doc_source"] == res["target_doc"]
                    ws1.cell(row=row, column=4,  value=wa["rank"])
                    ws1.cell(row=row, column=5,  value=wa["chunk_id"])
                    ws1.cell(row=row, column=6,  value=wa["doc_source"])
                    ws1.cell(row=row, column=7,  value="YES" if correct else "no")
                    ws1.cell(row=row, column=8,  value=wa["text_score_norm"])
                    ws1.cell(row=row, column=9,  value=wa["heading_score_norm"])
                    ws1.cell(row=row, column=10, value=wa["combined_score"])
                    fill = green_fill if correct else red_fill
                    for col in range(4, 11):
                        ws1.cell(row=row, column=col).fill = fill

                if i < len(res["without_ancestral"]):
                    wo      = res["without_ancestral"][i]
                    correct = wo["doc_source"] == res["target_doc"]
                    ws1.cell(row=row, column=11, value=wo["rank"])
                    ws1.cell(row=row, column=12, value=wo["chunk_id"])
                    ws1.cell(row=row, column=13, value=wo["doc_source"])
                    ws1.cell(row=row, column=14, value="YES" if correct else "no")
                    ws1.cell(row=row, column=15, value=wo["text_score_norm"])
                    fill = green_fill if correct else red_fill
                    for col in range(11, 16):
                        ws1.cell(row=row, column=col).fill = fill

                for col in range(1, 16):
                    ws1.cell(row=row, column=col).border = thin_border

            if n_rows > 1:
                for col in [1, 2, 3]:
                    ws1.merge_cells(
                        start_row=cur_row, start_column=col,
                        end_row=cur_row + n_rows - 1, end_column=col,
                    )
                    ws1.cell(row=cur_row, column=col).alignment = Alignment(
                        wrap_text=True, vertical="center", horizontal="center",
                    )

            cur_row += n_rows

        for col_letter, w in zip("ABCDEFGHIJKLMNO",
                                  [7, 54, 28, 6, 13, 30, 9, 11, 13, 11, 6, 13, 30, 9, 11]):
            ws1.column_dimensions[col_letter].width = w

        ws2 = wb.create_sheet("Precision Summary")

        sum_hdrs = [
            "Query ID", "Question (truncated)", "Target Document",
            f"P@{k} With Ancestral", f"P@{k} Without Ancestral",
            "Delta (With \u2212 Without)", "Result",
        ]
        for ci, h in enumerate(sum_hdrs, 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.font = hdr_font; c.fill = hdr_fill_blu; c.border = thin_border; c.alignment = ctr

        for ri, res in enumerate(all_results, 2):
            delta  = res["precision_delta"]
            result = "With Ancestral \u2714" if delta > 0 else ("Tie" if delta == 0 else "Without Ancestral")
            vals   = [
                res["query_id"],
                res["question"][:80] + ("..." if len(res["question"]) > 80 else ""),
                res["target_doc"],
                f"{res['precision_with']:.0%}",
                f"{res['precision_without']:.0%}",
                f"{delta:+.0%}",
                result,
            ]
            for ci, v in enumerate(vals, 1):
                c = ws2.cell(row=ri, column=ci, value=v)
                c.border    = thin_border
                c.alignment = Alignment(
                    wrap_text=True, vertical="center",
                    horizontal="left" if ci == 2 else "center",
                )
            row_fill = green_fill if delta > 0 else (red_fill if delta < 0 else gold_fill)
            for col in [4, 5, 6, 7]:
                ws2.cell(row=ri, column=col).fill = row_fill

        avg_row = len(all_results) + 3
        vals_avg = [
            "AVERAGE",
            f"{len(all_results)} queries across 3 documents",
            f"K = {k}",
            f"{avg_with:.0%}",
            f"{avg_without:.0%}",
            f"{avg_delta:+.0%}",
            "With Ancestral \u2714" if avg_delta > 0 else ("Tie" if avg_delta == 0 else "Without Ancestral"),
        ]
        for ci, v in enumerate(vals_avg, 1):
            c = ws2.cell(row=avg_row, column=ci, value=v)
            c.font      = Font(bold=True, size=11)
            c.border    = thin_border
            c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            c.fill      = green_fill if avg_delta > 0 else gold_fill

        for col_letter, w in zip("ABCDEFG", [10, 58, 32, 18, 22, 22, 22]):
            ws2.column_dimensions[col_letter].width = w
        for row in range(2, avg_row + 1):
            ws2.row_dimensions[row].height = 36

        excel_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(excel_path))
        print(f"\nExcel saved to: {excel_path}")

    finally:
        client.close()
        print("Done.")


if __name__ == "__main__":
    _config_path = Path(__file__).resolve().parent.parent / "config" / "config.json"
    with open(_config_path, "r", encoding="utf-8") as _f:
        _full = json.load(_f)
    _cfg = {**_full, **_full["eval_multidoc"]}
    run(_cfg)
