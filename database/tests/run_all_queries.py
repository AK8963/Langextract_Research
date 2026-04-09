"""
Automation script: generates queries from input JSON chunks, runs each query
through Weaviate with both --with_ancestral and --without_ancestral modes,
and writes all results to a single Excel file.

Usage:
    python run_all_queries.py
"""

import time
import warnings
import requests
import weaviate
from weaviate.classes.config import Configure, Property, DataType
from weaviate.classes.query import MetadataQuery, HybridFusion, TargetVectors, Filter
from weaviate.config import AdditionalConfig, Timeout
from weaviate.exceptions import WeaviateQueryError
import math
import unicodedata
import json
from pathlib import Path
import sys
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Ensure UTF-8 output on Windows
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ========================
# LOAD CONFIG
# ========================
_config_path = Path(__file__).resolve().parent.parent / "config" / "config.json"
with open(_config_path, "r", encoding="utf-8") as _cfg_f:
    _full_cfg = json.load(_cfg_f)
# Merge shared keys (weaviate, ollama) with the db-specific section
cfg = {**_full_cfg, **_full_cfg["db"]}

warnings.filterwarnings("ignore", category=DeprecationWarning, module="weaviate")

OLLAMA_HOST      = cfg["ollama"]["host"]
OLLAMA_DOCKER_EP = cfg["ollama"]["docker_endpoint"]
EMBED_MODEL      = cfg["ollama"]["embed_model"]
GENERATIVE_MODEL = cfg["ollama"]["generative_model"]
W_TEXT           = cfg["retrieval"]["w_text"]
W_HEADING        = cfg["retrieval"]["w_heading"]
TOP_K_RETRIEVE   = cfg["retrieval"]["top_k_retrieve"]
TOP_N            = cfg["retrieval"]["top_n"]
DATA_FILE        = cfg["data_file"]
QUERIES_MD_FILE  = cfg.get("queries_md_file", "queries/generated_queries.md")
EXCEL_OUTPUT_FILE = cfg.get("excel_output_file", "output/retrieval_results.xlsx")
METADATA_FILTERS  = cfg.get("metadata_filters", {})

BASE_DIR = Path(__file__).resolve().parent.parent

# ========================
# HELPER FUNCTIONS
# ========================
def _clean(s: str) -> str:
    if not s:
        return ""
    s = unicodedata.normalize("NFKC", s)
    s = "".join(c for c in s if unicodedata.category(c) not in ("Cc", "Cf") or c in "\t\n\r")
    return s.strip()

def build_heading_path(chunk_data: dict) -> str:
    anc = chunk_data.get("ancestral_headings", {})
    if not isinstance(anc, dict):
        return ""
    paths = []
    for heading, ancestors in anc.items():
        if isinstance(ancestors, list) and ancestors:
            paths.append(" > ".join(ancestors + [heading]))
        else:
            paths.append(heading)
    if paths:
        paths.sort(key=lambda p: p.count(" > "), reverse=True)
        return paths[0]
    return ""

def _embed_direct(text: str) -> list | None:
    try:
        r = requests.post(
            f"{OLLAMA_HOST}/api/embed",
            json={"model": EMBED_MODEL, "input": text or "document"},
            timeout=60,
        )
        r.raise_for_status()
        vec = r.json().get("embeddings", [[]])[0]
        if any(math.isnan(v) or math.isinf(v) for v in vec):
            return None
        return vec
    except Exception:
        return None


def _extract_metadata_props(chunk_data: dict, raw_text: str, data_file_path: str) -> dict:
    """Derive filterable metadata properties from a chunk for storage alongside the document."""
    doc_source = Path(data_file_path).stem

    # Determine deepest heading level present in this chunk's keys
    heading_level = "main"
    for key in chunk_data.keys():
        if key == "ancestral_headings":
            continue
        if "Sub Sub Sub" in key:
            heading_level = "sub_sub_sub"
            break
        elif "Sub Sub" in key:
            heading_level = "sub_sub"
        elif "Sub" in key and heading_level not in ("sub_sub",):
            heading_level = "sub"

    # top_section: the root ancestor (no parents) from ancestral_headings
    top_section = ""
    anc = chunk_data.get("ancestral_headings", {})
    if isinstance(anc, dict):
        for heading, ancestors in anc.items():
            if not ancestors:
                top_section = heading
                break
        if not top_section and anc:
            shallowest = min(anc.items(), key=lambda kv: len(kv[1]))
            top_section = shallowest[0]

    return {
        "doc_source":    doc_source,
        "heading_level": heading_level,
        "top_section":   top_section,
        "has_code":      "```" in raw_text,
        "has_images":    "![" in raw_text,
    }


def _build_filters(filter_cfg: dict):
    """Build a Weaviate Filter from the metadata_filters config block.
    Returns None when no filters are configured (→ no filtering applied)."""
    conditions = []
    val = filter_cfg.get("doc_source")
    if val:
        conditions.append(Filter.by_property("doc_source").equal(val))
    val = filter_cfg.get("heading_level")
    if val:
        conditions.append(Filter.by_property("heading_level").equal(val))
    val = filter_cfg.get("top_section")
    if val:
        conditions.append(Filter.by_property("top_section").equal(val))
    val = filter_cfg.get("has_code")
    if val is not None:
        conditions.append(Filter.by_property("has_code").equal(bool(val)))
    val = filter_cfg.get("has_images")
    if val is not None:
        conditions.append(Filter.by_property("has_images").equal(bool(val)))
    if not conditions:
        return None
    return conditions[0] if len(conditions) == 1 else Filter.all_of(conditions)


# ========================
# 1. LOAD DATA & GENERATE QUERIES VIA LLM
# ========================
json_path = Path(DATA_FILE)
if not json_path.exists():
    raise FileNotFoundError(f"Data file not found: {json_path}\nUpdate 'data_file' in config.json.")

print(f"Loading data from: {json_path}")
with open(json_path, "r", encoding="utf-8") as f:
    data = json.load(f)

queries_md_path = BASE_DIR / QUERIES_MD_FILE
queries_md_path.parent.mkdir(parents=True, exist_ok=True)

# Check if Excel output already exists — refuse to overwrite
excel_path = BASE_DIR / EXCEL_OUTPUT_FILE
if excel_path.exists():
    raise FileExistsError(
        f"Excel output already exists: {excel_path}\n"
        "Rename/delete it or change 'excel_output_file' in config.json."
    )


def generate_query_via_llm(chunk_id: str, chunk_text: str, headings: list,
                           ancestral_headings: dict, metadata: str = "") -> str:
    """Use local Ollama LLM to generate a natural human-like question for this chunk."""
    import re as _re
    import random

    # Clean the text: remove markdown image refs, table formatting noise
    clean_text = chunk_text
    clean_text = _re.sub(r'!\[\]\([^)]*\)', '', clean_text)
    clean_text = _re.sub(r'\|[-\s|]+\|', '', clean_text)
    clean_text = _re.sub(r'[#*]+', '', clean_text)
    clean_text = _re.sub(r'\n{3,}', '\n\n', clean_text)
    clean_text = clean_text.strip()

    text_snippet = clean_text[:1500] if len(clean_text) > 1500 else clean_text

    # Build metadata snippet
    meta_snippet = ""
    if metadata:
        meta_snippet = f"\nContext: {metadata[:500]}"

    # Pick a random style so questions don't all start the same way
    styles = [
        "Ask as a customer wanting to BUY or USE this — e.g. 'I want to ... but how does ... work with ...'",
        "Ask as a client COMPARING options — e.g. 'What is the difference between ... and ... and which is better for ...'",
        "Ask as a client who is CONFUSED — e.g. 'Why does ... need ... when ... already does ...'",
        "Ask as a new user needing HELP — e.g. 'How do I get started with ... if I already have ...'",
        "Ask as a client with a PROBLEM — e.g. 'What happens when ... goes wrong and how do I fix ...'",
        "Ask as someone EVALUATING the product — e.g. 'Is it true that ... supports ... and can it handle ...'",
    ]
    style = random.choice(styles)

    prompt = f"""You are a CLIENT or CUSTOMER who wants to know something about the product, service, or topic below. The person answering is the owner/expert. Write ONE natural question a client would ask.

Content:
{text_snippet}
{meta_snippet}

STYLE: {style}

RULES:
1. Ask about SPECIFIC details from the content — use actual names, numbers, features, steps, or facts
2. Sound like a real person talking — casual, direct, natural
3. NEVER start with "I noticed", "I see that", "I read that"
4. NEVER mention any heading, title, section, chapter, or document
5. NEVER say "described in", "mentioned in", "according to", "in the text", "in this section"
6. Ask as if you are having a CONVERSATION about the topic, not reading a document
7. Combine 2-3 different specific details into one complex question
8. The question should require deep understanding to answer — not just keyword matching
9. Output ONLY the question — nothing else, no prefix, no quotes

Question:"""

    try:
        resp = requests.post(
            f"{OLLAMA_HOST}/api/generate",
            json={
                "model": GENERATIVE_MODEL,
                "prompt": prompt,
                "stream": False,
                "options": {"temperature": 0.9, "num_predict": 200},
            },
            timeout=180,
        )
        resp.raise_for_status()
        result = resp.json()
        answer = result.get("response", "").strip()
        print(f"    [LLM raw response length: {len(answer)} chars]")
        if not answer:
            print(f"    [LLM returned EMPTY response for {chunk_id}]")

        # Clean up — take first meaningful question line
        for line in answer.split("\n"):
            line = line.strip().strip('"').strip("'").strip()
            if not line:
                continue
            # Skip meta-prefixes
            low = line.lower()
            if low.startswith("here") or low.startswith("question") or low.startswith("sure"):
                continue
            # Strip leading numbering like "1." or "1)"
            line = _re.sub(r'^\d+[\.\)]\s*', '', line).strip()
            if not line:
                continue
            if "?" in line:
                return line
            if len(line) > 30:
                return line + ("?" if not line.endswith("?") else "")
        if answer:
            first_line = answer.split("\n")[0].strip().strip('"').strip("'").strip()
            if first_line:
                return first_line + ("?" if not first_line.endswith("?") else "")
    except Exception as e:
        print(f"  LLM query generation failed for {chunk_id}: {e}")

    # ── Content-aware fallback (no heading names!) ──
    print(f"    [Using text-based fallback for {chunk_id}]")

    text_sample = _re.sub(r'[#*|\[\]`]', '', chunk_text[:600]).strip()

    # Extract specific nouns/verbs from text
    stop_words = {"this", "that", "with", "from", "have", "been", "their", "there",
                  "were", "which", "about", "would", "could", "should", "these",
                  "those", "through", "between", "after", "before", "under", "above",
                  "into", "more", "than", "also", "very", "just", "only", "report",
                  "heading", "section", "document", "summary", "overview", "analysis",
                  "following", "using", "based", "including", "provide", "information"}
    words = [w for w in _re.findall(r'[a-zA-Z]{5,}', text_sample)
             if w.lower() not in stop_words]
    seen = set()
    key_words = []
    for w in words:
        wl = w.lower()
        if wl not in seen:
            seen.add(wl)
            key_words.append(w)
        if len(key_words) >= 4:
            break

    # Build fallback from actual content keywords
    if len(key_words) >= 3:
        return (f"If I need to work with {key_words[0].lower()} and {key_words[1].lower()}, "
                f"how does {key_words[2].lower()} factor into that, and what should I watch out for?")
    elif len(key_words) >= 2:
        return (f"How does {key_words[0].lower()} connect to {key_words[1].lower()}, "
                f"and what's the right way to handle both?")
    else:
        sentences = [s.strip() for s in text_sample.split('.') if len(s.strip()) > 20]
        if sentences:
            return f"What's the best approach when dealing with {sentences[0][:80].lower().rstrip()}?"
        return "What are the key steps here and what should I focus on first?"


chunk_queries = []  # list of (chunk_id, query_text, chunk_context_snippet)

# If MD file already exists, parse queries from it instead of regenerating
if queries_md_path.exists():
    print(f"\nQueries MD file already exists: {queries_md_path}")
    print("Loading queries from existing file (delete it to regenerate via LLM)...")

    import re
    md_content = queries_md_path.read_text(encoding="utf-8")
    # Parse "## Query N — chunk_idX" blocks and extract the query line
    query_blocks = re.split(r"^## Query \d+ — ", md_content, flags=re.MULTILINE)[1:]
    for block in query_blocks:
        lines = block.strip().split("\n")
        # First line is the chunk_id
        cid = lines[0].strip()
        # Find the query — either a **Query:** line or plain text after chunk_id
        qtxt = ""
        for line in lines[1:]:
            line = line.strip()
            if line.startswith("**Query:**"):
                qtxt = line.replace("**Query:**", "").strip()
                break
            elif line and not line.startswith("---") and not line.startswith("**Context"):
                qtxt = line
                break
        if not qtxt:
            continue
        # Find matching chunk context from data
        context_snippet = ""
        for item in data:
            for key, val in item.items():
                if key == cid and isinstance(val, dict):
                    text = item.get("Text", "")
                    context_snippet = (text[:300].replace("\n", " ").strip() + "...") if len(text) > 300 else text.replace("\n", " ").strip()
                    break
        chunk_queries.append((cid, qtxt, context_snippet))

    print(f"Loaded {len(chunk_queries)} queries from MD file.")

else:
    MAX_CHUNKS = 10
    print(f"\nGenerating queries via LLM (max {MAX_CHUNKS} chunks)...")
    for item in data:
        for chunk_id, chunk_data in item.items():
            if len(chunk_queries) >= MAX_CHUNKS:
                break
            if chunk_id in ("Text", "Metadata", "ancestral_headings") or not isinstance(chunk_data, dict):
                continue
            if not chunk_id.startswith("chunk_id"):
                continue

            text = item.get("Text", "")
            metadata = item.get("Metadata", "")
            headings = []
            for key, value in chunk_data.items():
                if key == "ancestral_headings":
                    continue
                if value and isinstance(value, str):
                    # Strip trailing dots and whitespace from heading values
                    clean_heading = value.replace(".....", "").rstrip(".").strip()
                    if clean_heading:
                        headings.append(clean_heading)

            ancestral_headings = chunk_data.get("ancestral_headings", {})

            print(f"  Generating query for {chunk_id}...")
            query_text = generate_query_via_llm(chunk_id, text, headings, ancestral_headings, metadata)
            print(f"    → {query_text[:100]}...")

            context_snippet = (text[:300].replace("\n", " ").strip() + "...") if len(text) > 300 else text.replace("\n", " ").strip()
            chunk_queries.append((chunk_id, query_text, context_snippet))
        if len(chunk_queries) >= MAX_CHUNKS:
            break

    # Write queries to MD file (path from config) — queries only
    with open(queries_md_path, "w", encoding="utf-8") as qf:
        qf.write("# LLM-Generated Queries\n\n")
        qf.write(f"Source: `{json_path.name}`\n")
        qf.write(f"Model: `{GENERATIVE_MODEL}`\n\n")
        for i, (cid, qtxt, ctx) in enumerate(chunk_queries, 1):
            qf.write(f"## Query {i} — {cid}\n\n")
            qf.write(f"{qtxt}\n\n---\n\n")

    print(f"\nGenerated {len(chunk_queries)} queries → {queries_md_path}")


# ========================
# 2. CONNECT TO WEAVIATE & INSERT DATA (once)
# ========================
print("\nConnecting to Weaviate...")
client = weaviate.connect_to_custom(
    http_host=cfg["weaviate"]["http_host"],
    http_port=cfg["weaviate"]["http_port"],
    http_secure=False,
    grpc_host=cfg["weaviate"]["grpc_host"],
    grpc_port=cfg["weaviate"]["grpc_port"],
    grpc_secure=False,
    skip_init_checks=True,
    additional_config=AdditionalConfig(
        timeout=Timeout(init=30, query=120, insert=120)
    ),
)

try:
    # Recreate collection
    if client.collections.exists("LangExtractDocs"):
        client.collections.delete("LangExtractDocs")

    client.collections.create(
        name="LangExtractDocs",
        vector_config=[
            Configure.Vectors.text2vec_ollama(
                name="text_vector",
                source_properties=["text"],
                api_endpoint=OLLAMA_DOCKER_EP,
                model=EMBED_MODEL,
            ),
            Configure.Vectors.text2vec_ollama(
                name="heading_vector",
                source_properties=["heading_path"],
                api_endpoint=OLLAMA_DOCKER_EP,
                model=EMBED_MODEL,
            ),
        ],
        generative_config=Configure.Generative.ollama(
            api_endpoint=OLLAMA_DOCKER_EP,
            model=GENERATIVE_MODEL,
        ),
        properties=[
            Property(name="text", data_type=DataType.TEXT),
            Property(name="heading_path", data_type=DataType.TEXT),
            Property(name="main_heading", data_type=DataType.TEXT, skip_vectorization=True),
            Property(name="sub_headings", data_type=DataType.TEXT_ARRAY, skip_vectorization=True),
            Property(name="ancestral_headings", data_type=DataType.TEXT, skip_vectorization=True),
            Property(name="metadata_str", data_type=DataType.TEXT, skip_vectorization=True),
            Property(name="chunk_id", data_type=DataType.TEXT, skip_vectorization=True),
            # Filterable metadata properties
            Property(name="doc_source", data_type=DataType.TEXT, skip_vectorization=True),
            Property(name="heading_level", data_type=DataType.TEXT, skip_vectorization=True),
            Property(name="top_section", data_type=DataType.TEXT, skip_vectorization=True),
            Property(name="has_code", data_type=DataType.BOOL, skip_vectorization=True),
            Property(name="has_images", data_type=DataType.BOOL, skip_vectorization=True),
        ],
    )
    print("Collection 'LangExtractDocs' created.")

    collection = client.collections.use("LangExtractDocs")

    # Insert all chunks
    for item in data:
        for chunk_id, chunk_data in item.items():
            if chunk_id in ("Text", "Metadata", "ancestral_headings") or not isinstance(chunk_data, dict):
                continue
            if not chunk_id.startswith("chunk_id"):
                continue

            text = item.get("Text", "")
            metadata_str = item.get("Metadata", "")
            main_heading = ""
            sub_headings = []
            for key, value in chunk_data.items():
                if key == "ancestral_headings":
                    continue
                if "Main Heading" in key:
                    main_heading = value
                elif "Sub Heading" in key or "Sub Sub Heading" in key:
                    sub_headings.append(value)

            ancestral_headings_str = json.dumps(chunk_data.get("ancestral_headings", {}))
            heading_path = build_heading_path(chunk_data)

            text = _clean(str(text) if text is not None else "")
            heading_path = _clean(str(heading_path) if heading_path is not None else "")

            if not heading_path:
                if main_heading and main_heading.strip():
                    heading_path = _clean(main_heading)
                elif text:
                    heading_path = text[:80]
                else:
                    print(f"Skipping chunk {chunk_id}: no text or heading available.")
                    continue

            if not heading_path or not heading_path.strip():
                heading_path = f"Section: {chunk_id}"

            _meta_props = _extract_metadata_props(chunk_data, text, DATA_FILE)
            _insert_obj = {
                "text": text,
                "heading_path": heading_path,
                "main_heading": main_heading,
                "sub_headings": sub_headings,
                "ancestral_headings": ancestral_headings_str,
                "metadata_str": metadata_str,
                "chunk_id": chunk_id,
                **_meta_props,
            }

            safe_text = _insert_obj["text"]
            safe_hp = _insert_obj["heading_path"]

            inserted = False
            for insert_attempt in range(3):
                try:
                    collection.data.insert(_insert_obj)
                    inserted = True
                    break
                except Exception as e:
                    err_msg = str(e)
                    if insert_attempt == 0:
                        time.sleep(1)
                        _insert_obj = {**_insert_obj, "text": _insert_obj["text"][:512], "heading_path": _insert_obj["heading_path"][:100]}
                    elif insert_attempt == 1:
                        time.sleep(2)
                        safe_text = "".join(c for c in _insert_obj["text"][:256] if ord(c) < 128).strip() or "document content"
                        safe_hp = "".join(c for c in _insert_obj["heading_path"] if ord(c) < 128).strip() or "document section"
                        _insert_obj = {**_insert_obj, "text": safe_text, "heading_path": safe_hp}
                    else:
                        t_vec = _embed_direct(safe_text) or _embed_direct("document content")
                        h_vec = _embed_direct(safe_hp) or _embed_direct("document section")
                        if t_vec and h_vec:
                            try:
                                collection.data.insert(_insert_obj, vector={"text_vector": t_vec, "heading_vector": h_vec})
                                inserted = True
                            except Exception:
                                pass

            if not inserted:
                print(f"Failed to insert {chunk_id}")

    print("Data inserted successfully.\n")

    # Wait for Weaviate to finish vectorizing all chunks before querying
    print("Waiting for vectorization to complete...")
    time.sleep(10)

    # Build metadata filter once — shared across all query runs
    _active_filters = _build_filters(METADATA_FILTERS)
    if _active_filters is not None:
        print(f"Metadata filters active: { {k: v for k, v in METADATA_FILTERS.items() if v not in (None, '')} }")

    # ========================
    # 3. RUN QUERIES — both modes
    # ========================
    def run_retrieval(query_text: str, include_ancestral: bool):
        """Run a single query and return a list of result dicts."""
        # Text search
        search_text = None
        for attempt in range(3):
            try:
                search_text = collection.query.hybrid(
                    query=query_text,
                    target_vector=TargetVectors.sum(["text_vector"]),
                    query_properties=["text"],
                    alpha=0.7,
                    fusion_type=HybridFusion.RELATIVE_SCORE,
                    limit=TOP_K_RETRIEVE,
                    filters=_active_filters,
                    return_metadata=MetadataQuery(score=True),
                )
                break
            except WeaviateQueryError:
                if attempt < 2:
                    time.sleep(2 * (attempt + 1))

        # Heading search
        search_heading = None
        for attempt in range(3):
            try:
                search_heading = collection.query.hybrid(
                    query=query_text,
                    target_vector=TargetVectors.sum(["heading_vector"]),
                    query_properties=["heading_path"],
                    alpha=0.7,
                    fusion_type=HybridFusion.RELATIVE_SCORE,
                    limit=TOP_K_RETRIEVE,
                    filters=_active_filters,
                    return_metadata=MetadataQuery(score=True),
                )
                break
            except WeaviateQueryError:
                if attempt < 2:
                    time.sleep(2 * (attempt + 1))

        if (not search_text or not getattr(search_text, 'objects', None)) and \
           (not search_heading or not getattr(search_heading, 'objects', None)):
            return []

        # Build score maps
        text_scores = {}
        heading_scores = {}
        if search_text and getattr(search_text, 'objects', None):
            for o in search_text.objects:
                cid = o.properties.get('chunk_id', '')
                s = o.metadata.score if o.metadata and o.metadata.score is not None else 0.0
                text_scores[cid] = max(text_scores.get(cid, 0.0), s)
        if search_heading and getattr(search_heading, 'objects', None):
            for o in search_heading.objects:
                cid = o.properties.get('chunk_id', '')
                s = o.metadata.score if o.metadata and o.metadata.score is not None else 0.0
                heading_scores[cid] = max(heading_scores.get(cid, 0.0), s)

        all_chunk_ids = set(list(text_scores.keys()) + list(heading_scores.keys()))
        max_text = max(text_scores.values()) if text_scores else 1.0
        max_heading = max(heading_scores.values()) if heading_scores else 1.0
        if max_text == 0:
            max_text = 1.0
        if max_heading == 0:
            max_heading = 1.0

        def find_obj(cid):
            for src in (search_text, search_heading):
                if not src or not getattr(src, 'objects', None):
                    continue
                for o in src.objects:
                    if o.properties.get('chunk_id', '') == cid:
                        return o
            return None

        combined_list = []
        for cid in all_chunk_ids:
            t = text_scores.get(cid, 0.0) / max_text
            h = heading_scores.get(cid, 0.0) / max_heading
            combined = W_TEXT * t + W_HEADING * h
            obj = find_obj(cid)
            combined_list.append((cid, t, h, combined, obj))

        if include_ancestral:
            combined_list.sort(key=lambda x: x[3], reverse=True)
        else:
            combined_list.sort(key=lambda x: x[1], reverse=True)

        selected = combined_list[:TOP_N]
        records = []
        for rank, (cid, t_score, h_score, comb, wobj) in enumerate(selected, 1):
            if wobj is None:
                continue
            if include_ancestral:
                records.append({
                    "rank": rank,
                    "chunk_id": cid,
                    "text_score": round(t_score, 6),
                    "ancestral_score": round(h_score, 6),
                    "combined_score": round(comb, 6),
                })
            else:
                records.append({
                    "rank": rank,
                    "chunk_id": cid,
                    "retrieval_score": round(t_score, 6),
                })
        return records

    # Collect all results
    all_results = []  # list of dicts for Excel rows

    total = len(chunk_queries)
    for idx, (source_chunk_id, query_text, context_snippet) in enumerate(chunk_queries, 1):
        print(f"[{idx}/{total}] Running query for {source_chunk_id} ...")

        with_anc = run_retrieval(query_text, include_ancestral=True)
        without_anc = run_retrieval(query_text, include_ancestral=False)

        all_results.append({
            "source_chunk_id": source_chunk_id,
            "question": query_text,
            "chunk_context": context_snippet,
            "with_ancestral": with_anc,
            "without_ancestral": without_anc,
        })
        print(f"  with_ancestral: {len(with_anc)} results, without_ancestral: {len(without_anc)} results")

    # ========================
    # 4. WRITE EXCEL
    # ========================
    wb = Workbook()
    ws = wb.active
    ws.title = "Retrieval Results"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    sub_header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    sub_header_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    wrap_align = Alignment(wrap_text=True, vertical="top")

    # Highlight fills for rank comparison
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")   # ancestral better
    red_fill = PatternFill(start_color="FF5757", end_color="FF5757", fill_type="solid")     # ancestral worse
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # same / without-ancestral marker

    # Headers row 1 — main groups
    # Columns:
    # A: Question
    # B: Source Chunk ID
    # C: Chunk Context
    # --- With Ancestral-Heading (D-H) ---
    # D: rank, E: chunk_id, F: text_score, G: ancestral_score, H: combined_score
    # --- Without Ancestral-Heading (I-K) ---
    # I: rank, J: chunk_id, K: retrieval_score

    # Row 1: merged group headers
    ws.merge_cells("A1:C1")
    ws["A1"] = "Query Info"
    ws.merge_cells("D1:H1")
    ws["D1"] = "With Ancestral-Heading"
    ws.merge_cells("I1:K1")
    ws["I1"] = "Without Ancestral-Heading"

    for cell in [ws["A1"], ws["D1"], ws["I1"]]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Row 2: sub-headers
    sub_headers = [
        "Question", "Source Chunk ID", "Chunk Context",
        "rank", "chunk_id", "text_score", "ancestral_score", "combined_score",
        "rank", "chunk_id", "retrieval_score",
    ]
    for col_idx, hdr in enumerate(sub_headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=hdr)
        cell.font = sub_header_font
        cell.fill = sub_header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data — each query gets TOP_N rows (one per result rank)
    current_row = 3
    for res in all_results:
        n_rows = max(len(res["with_ancestral"]), len(res["without_ancestral"]), 1)
        source_cid = res["source_chunk_id"]

        # Find the rank of the source chunk in each mode (None if not in top N)
        anc_rank = None
        for r in res["with_ancestral"]:
            if r["chunk_id"] == source_cid:
                anc_rank = r["rank"]
                break
        no_anc_rank = None
        for r in res["without_ancestral"]:
            if r["chunk_id"] == source_cid:
                no_anc_rank = r["rank"]
                break

        # Determine highlight for the source chunk row in "With Ancestral" side
        # Green = ancestral rank is better (lower number) than without-ancestral rank
        # Red   = ancestral rank is worse (higher number) than without-ancestral rank
        # Yellow = same rank, or one/both modes don't have the source chunk in top N
        if anc_rank is not None and no_anc_rank is not None:
            if anc_rank < no_anc_rank:
                anc_highlight = green_fill
            elif anc_rank > no_anc_rank:
                anc_highlight = red_fill
            else:
                anc_highlight = yellow_fill
        elif anc_rank is not None and no_anc_rank is None:
            # Present in ancestral but not in without — ancestral is better
            anc_highlight = green_fill
        elif anc_rank is None and no_anc_rank is not None:
            # Not in ancestral but present in without — ancestral is worse
            anc_highlight = red_fill
        else:
            anc_highlight = yellow_fill

        # Write query info on first row only, then merge across all result rows
        ws.cell(row=current_row, column=1, value=res["question"]).alignment = wrap_align
        ws.cell(row=current_row, column=2, value=res["source_chunk_id"]).alignment = wrap_align
        ws.cell(row=current_row, column=3, value=res["chunk_context"]).alignment = wrap_align

        for i in range(n_rows):
            row = current_row + i

            # With Ancestral columns (D-H)
            if i < len(res["with_ancestral"]):
                wa = res["with_ancestral"][i]
                ws.cell(row=row, column=4, value=wa["rank"])
                ws.cell(row=row, column=5, value=wa["chunk_id"])
                ws.cell(row=row, column=6, value=wa["text_score"])
                ws.cell(row=row, column=7, value=wa["ancestral_score"])
                ws.cell(row=row, column=8, value=wa["combined_score"])

                # Highlight the row where chunk_id matches the source chunk
                if wa["chunk_id"] == source_cid:
                    for col in range(4, 9):  # columns D-H
                        ws.cell(row=row, column=col).fill = anc_highlight

            # Without Ancestral columns (I-K)
            if i < len(res["without_ancestral"]):
                woa = res["without_ancestral"][i]
                ws.cell(row=row, column=9, value=woa["rank"])
                ws.cell(row=row, column=10, value=woa["chunk_id"])
                ws.cell(row=row, column=11, value=woa["retrieval_score"])

                # Always yellow for the source chunk row in Without Ancestral
                if woa["chunk_id"] == source_cid:
                    for col in range(9, 12):  # columns I-K
                        ws.cell(row=row, column=col).fill = yellow_fill

            # Borders on all columns
            for col in range(1, 12):
                ws.cell(row=row, column=col).border = thin_border

        # Merge query-info cells (A, B, C) across all result rows for this query
        if n_rows > 1:
            for col in [1, 2, 3]:
                ws.merge_cells(start_row=current_row, start_column=col,
                               end_row=current_row + n_rows - 1, end_column=col)
                ws.cell(row=current_row, column=col).alignment = Alignment(
                    wrap_text=True, vertical="center")

        current_row += n_rows

    # Column widths
    col_widths = {"A": 50, "B": 15, "C": 60,
                  "D": 8, "E": 15, "F": 12, "G": 15, "H": 15,
                  "I": 8, "J": 15, "K": 15}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Save
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(excel_path))
    print(f"\nExcel saved to: {excel_path}")

finally:
    client.close()
    print("Done.")
