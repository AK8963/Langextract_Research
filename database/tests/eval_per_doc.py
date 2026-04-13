"""
Per-Document Title Query Evaluation — Proves ancestral-heading superiority.

For every JSON in Task2/output/ (auto-discovered, excludes combined/notes files):
  - Derives the query from the ROOT heading in ancestral_headings (the doc title).
  - Inserts ALL chunks from ALL docs into a single shared Weaviate index.
  - Runs retrieval with and without ancestral headings for each doc's title query.
  - Computes 28 LG-mapped metrics + derived metrics.
  - Writes a 2-sheet Excel file:
      Sheet 1 "Detailed Results" – one row per (doc × retrieved rank)
      Sheet 2 "Summary"         – one row per doc with all metrics + colored footers

Chunk IDs are prefixed "{file_stem}::{chunk_id_key}" to guarantee uniqueness
across the multi-document index.

Intended to be called from main.py:
    from eval_per_doc import run
    run(cfg)
"""

import time
import warnings
import math
import unicodedata
import json
from pathlib import Path
from collections import Counter

import requests
import weaviate
from weaviate.classes.config import Configure, Property, DataType
from weaviate.classes.query import MetadataQuery, HybridFusion, TargetVectors
from weaviate.config import AdditionalConfig, Timeout
from weaviate.exceptions import WeaviateQueryError
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=DeprecationWarning, module="weaviate")


# ========================
# PURE HELPERS (no config dependency)
# ========================
def _clean(s: str) -> str:
    if not s:
        return ""
    s = unicodedata.normalize("NFKC", s)
    s = "".join(c for c in s if unicodedata.category(c) not in ("Cc", "Cf") or c in "\t\n\r")
    return s.strip()


def build_heading_path(chunk_data: dict) -> str:
    """Build the richest heading path from ancestral_headings for the heading vector."""
    anc = chunk_data.get("ancestral_headings", {})
    if not isinstance(anc, dict):
        return ""
    paths = []
    for heading, ancestors in anc.items():
        if isinstance(ancestors, list) and ancestors:
            paths.append(" > ".join(ancestors + [heading]))
        else:
            paths.append(heading)
    if paths:
        paths.sort(key=lambda p: p.count(" > "), reverse=True)
        return paths[0]
    return ""


def get_doc_source(chunk_data: dict) -> str:
    """
    Return the root document name from ancestral_headings.
    The root heading is the one whose ancestors list is empty (depth = 0).
    Falls back to the first ancestor if no root is found.
    """
    anc = chunk_data.get("ancestral_headings", {})
    if not isinstance(anc, dict) or not anc:
        return "UNKNOWN"
    for heading, ancestors in anc.items():
        if isinstance(ancestors, list) and len(ancestors) == 0:
            return heading
    for heading, ancestors in anc.items():
        if isinstance(ancestors, list) and ancestors:
            return ancestors[0]
    return "UNKNOWN"


def _embed_direct(text: str, ollama_host: str, embed_model: str):
    try:
        r = requests.post(
            f"{ollama_host}/api/embed",
            json={"model": embed_model, "input": text or "document"},
            timeout=60,
        )
        r.raise_for_status()
        vec = r.json().get("embeddings", [[]])[0]
        if any(math.isnan(v) or math.isinf(v) for v in vec):
            return None
        return vec
    except Exception:
        return None


# ========================
# MAIN ENTRY POINT
# ========================
def run(cfg: dict) -> None:
    """
    Execute the per-document title query evaluation.

    Args:
        cfg: Parsed contents of eval_per_doc_config.json (or equivalent dict).
    """
    BASE_DIR = Path(__file__).resolve().parent.parent  # database/ — for output/ paths

    # ── Unpack config ────────────────────────────────────────────────────
    OLLAMA_HOST      = cfg["ollama"]["host"]
    OLLAMA_DOCKER_EP = cfg["ollama"]["docker_endpoint"]
    EMBED_MODEL      = cfg["ollama"]["embed_model"]
    GENERATIVE_MODEL = cfg["ollama"]["generative_model"]

    W_TEXT         = cfg["retrieval"]["w_text"]
    W_HEADING      = cfg["retrieval"]["w_heading"]
    TOP_K_RETRIEVE = cfg["retrieval"]["top_k_retrieve"]
    TOP_N          = cfg["retrieval"]["top_n"]
    ALPHA          = cfg["retrieval"].get("alpha", 0.5)

    TASK2_OUTPUT_DIR = Path(cfg["task2_output_dir"])
    EXCLUDE_FILES    = set(cfg.get("exclude_files", []))
    EXCEL_OUTPUT     = cfg.get("excel_output_file", "output/per_doc_title_query_results.xlsx")
    excel_path       = BASE_DIR / EXCEL_OUTPUT
    QUESTIONS_FILE   = BASE_DIR / cfg.get("questions_file", "tests/questions.json")

    # Bind config-dependent embed helper as a closure
    def embed_direct(text: str):
        return _embed_direct(text, OLLAMA_HOST, EMBED_MODEL)

    # ========================
    # DISCOVER & LOAD ALL DOCS
    # ========================
    if not TASK2_OUTPUT_DIR.exists():
        raise FileNotFoundError(f"Task2 output directory not found: {TASK2_OUTPUT_DIR}")

    json_files = sorted(
        f for f in TASK2_OUTPUT_DIR.glob("*.json")
        if f.name not in EXCLUDE_FILES
    )

    if not json_files:
        raise ValueError(f"No JSON files found in {TASK2_OUTPUT_DIR} (after excluding {EXCLUDE_FILES})")

    print(f"Discovered {len(json_files)} document files in {TASK2_OUTPUT_DIR}")

    doc_meta: list[dict] = []
    all_inserts: list[tuple[str, dict, dict, str]] = []

    for json_file in json_files:
        try:
            with open(json_file, "r", encoding="utf-8") as fh:
                data = json.load(fh)
        except Exception as e:
            print(f"  [WARN] Could not load {json_file.name}: {e}")
            continue

        if not isinstance(data, list) or not data:
            print(f"  [WARN] {json_file.name}: expected a non-empty list, skipping")
            continue

        stem = json_file.stem
        gt_ids: set[str] = set()
        root_heading = "UNKNOWN"

        for item in data:
            for chunk_key, chunk_data in item.items():
                if not (chunk_key.startswith("chunk_id") and isinstance(chunk_data, dict)):
                    continue
                prefixed = f"{stem}::{chunk_key}"
                gt_ids.add(prefixed)
                src = get_doc_source(chunk_data)
                if src != "UNKNOWN" and root_heading == "UNKNOWN":
                    root_heading = src
                all_inserts.append((prefixed, item, chunk_data, src))

        if not gt_ids:
            print(f"  [WARN] {json_file.name}: no chunk_id entries found, skipping")
            continue

        doc_meta.append({
            "file_stem":    stem,
            "root_heading": root_heading,
            "gt_chunk_ids": gt_ids,
            "gt_count":     len(gt_ids),
        })

    print(f"Loaded {len(doc_meta)} documents, {len(all_inserts)} total chunks")

    # ── Apply questions.json overrides ───────────────────────────────────────
    questions_data: dict = {}
    if QUESTIONS_FILE.exists():
        with open(QUESTIONS_FILE, "r", encoding="utf-8") as fh:
            questions_data = json.load(fh)
        print(f"Loaded questions from {QUESTIONS_FILE} ({len(questions_data)} entries)")
    else:
        print(f"[INFO] No questions file found at {QUESTIONS_FILE}; using auto-derived title queries only")

    for meta in doc_meta:
        stem = meta["file_stem"]
        if stem in questions_data:
            qentry = questions_data[stem]
            title_q = str(qentry.get("title_query", "")).strip()
            if title_q and title_q != "UNKNOWN":
                meta["root_heading"] = title_q
            meta["extra_questions"] = [
                str(q).strip() for q in qentry.get("extra_questions", []) if str(q).strip()
            ]
        else:
            meta["extra_questions"] = []

    # Warn about colliding root headings
    _root_heading_counter = Counter(m["root_heading"] for m in doc_meta)
    _collisions = {h: c for h, c in _root_heading_counter.items() if c > 1}
    if _collisions:
        print(f"[WARN] {len(_collisions)} root heading(s) shared by multiple files — "
              "doc_chunks_in_topN metrics will use chunk_id ground-truth sets for accuracy:")
        for h, c in _collisions.items():
            print(f"  '{h}' appears in {c} files")
    print()

    chunk_doc_map: dict[str, str] = {
        prefixed: doc_source
        for prefixed, _item, _chunk_data, doc_source in all_inserts
    }

    # ── Pre-flight ────────────────────────────────────────────────────────
    if excel_path.exists():
        raise FileExistsError(
            f"Excel output already exists: {excel_path}\n"
            "Rename or delete it before re-running."
        )

    # ========================
    # CONNECT TO WEAVIATE & INSERT ALL CHUNKS
    # ========================
    print("Connecting to Weaviate...")
    client = weaviate.connect_to_custom(
        http_host=cfg["weaviate"]["http_host"],
        http_port=cfg["weaviate"]["http_port"],
        http_secure=False,
        grpc_host=cfg["weaviate"]["grpc_host"],
        grpc_port=cfg["weaviate"]["grpc_port"],
        grpc_secure=False,
        skip_init_checks=True,
        additional_config=AdditionalConfig(timeout=Timeout(init=30, query=120, insert=120)),
    )

    try:
        if client.collections.exists("LangExtractEval"):
            client.collections.delete("LangExtractEval")

        client.collections.create(
            name="LangExtractEval",
            vector_config=[
                Configure.Vectors.text2vec_ollama(
                    name="text_vector",
                    source_properties=["text"],
                    api_endpoint=OLLAMA_DOCKER_EP,
                    model=EMBED_MODEL,
                ),
                Configure.Vectors.text2vec_ollama(
                    name="heading_vector",
                    source_properties=["heading_path"],
                    api_endpoint=OLLAMA_DOCKER_EP,
                    model=EMBED_MODEL,
                ),
            ],
            generative_config=Configure.Generative.ollama(
                api_endpoint=OLLAMA_DOCKER_EP,
                model=GENERATIVE_MODEL,
            ),
            properties=[
                Property(name="text",               data_type=DataType.TEXT),
                Property(name="heading_path",       data_type=DataType.TEXT),
                Property(name="chunk_id",           data_type=DataType.TEXT, skip_vectorization=True),
                Property(name="doc_source",         data_type=DataType.TEXT, skip_vectorization=True),
                Property(name="ancestral_headings", data_type=DataType.TEXT, skip_vectorization=True),
            ],
        )
        print("Collection 'LangExtractEval' created.")

        collection = client.collections.use("LangExtractEval")

        inserted_total = 0
        skipped_total  = 0

        for prefixed, item, chunk_data, doc_source in all_inserts:
            raw_text     = item.get("Text", "")
            text         = _clean(str(raw_text) if raw_text is not None else "")
            heading_path = _clean(build_heading_path(chunk_data))
            anc_str      = json.dumps(chunk_data.get("ancestral_headings", {}))

            if not heading_path:
                heading_path = doc_source or f"Section: {prefixed}"

            obj = {
                "text":               text,
                "heading_path":       heading_path,
                "chunk_id":           prefixed,
                "doc_source":         doc_source,
                "ancestral_headings": anc_str,
            }

            inserted = False
            for attempt in range(3):
                try:
                    collection.data.insert(obj)
                    inserted = True
                    break
                except Exception:
                    if attempt == 0:
                        time.sleep(1)
                        obj = {**obj, "text": obj["text"][:512], "heading_path": obj["heading_path"][:100]}
                    elif attempt == 1:
                        time.sleep(2)
                        safe_t = "".join(c for c in obj["text"][:256] if ord(c) < 128).strip() or "document content"
                        safe_h = "".join(c for c in obj["heading_path"] if ord(c) < 128).strip() or "document section"
                        obj = {**obj, "text": safe_t, "heading_path": safe_h}
                    else:
                        t_vec = embed_direct(obj["text"]) or embed_direct("document content")
                        h_vec = embed_direct(obj["heading_path"]) or embed_direct("document section")
                        if t_vec and h_vec:
                            try:
                                collection.data.insert(obj, vector={"text_vector": t_vec, "heading_vector": h_vec})
                                inserted = True
                            except Exception:
                                pass

            if inserted:
                inserted_total += 1
            else:
                print(f"  [WARN] Could not insert {prefixed}")
                skipped_total += 1

        print(f"Inserted {inserted_total}/{len(all_inserts)} chunks ({skipped_total} skipped). Waiting for vectorization...")
        time.sleep(10)

        # ========================
        # RETRIEVAL FUNCTION
        # ========================
        def run_retrieval(query_text: str, include_ancestral: bool) -> list[dict]:
            search_text = None
            for attempt in range(3):
                try:
                    search_text = collection.query.hybrid(
                        query=query_text,
                        target_vector=TargetVectors.sum(["text_vector"]),
                        query_properties=["text"],
                        alpha=ALPHA,
                        fusion_type=HybridFusion.RELATIVE_SCORE,
                        limit=TOP_K_RETRIEVE,
                        return_metadata=MetadataQuery(score=True),
                    )
                    break
                except WeaviateQueryError:
                    if attempt < 2:
                        time.sleep(2 * (attempt + 1))

            search_heading = None
            if include_ancestral:
                for attempt in range(3):
                    try:
                        search_heading = collection.query.hybrid(
                            query=query_text,
                            target_vector=TargetVectors.sum(["heading_vector"]),
                            query_properties=["heading_path"],
                            alpha=ALPHA,
                            fusion_type=HybridFusion.RELATIVE_SCORE,
                            limit=TOP_K_RETRIEVE,
                            return_metadata=MetadataQuery(score=True),
                        )
                        break
                    except WeaviateQueryError:
                        if attempt < 2:
                            time.sleep(2 * (attempt + 1))

            if not search_text or not getattr(search_text, "objects", None):
                return []

            text_scores:    dict[str, float] = {}
            heading_scores: dict[str, float] = {}
            obj_map:        dict[str, object] = {}

            for o in search_text.objects:
                cid = o.properties.get("chunk_id", "")
                s   = o.metadata.score if (o.metadata and o.metadata.score is not None) else 0.0
                text_scores[cid] = max(text_scores.get(cid, 0.0), s)
                obj_map[cid] = o

            if search_heading and getattr(search_heading, "objects", None):
                for o in search_heading.objects:
                    cid = o.properties.get("chunk_id", "")
                    s   = o.metadata.score if (o.metadata and o.metadata.score is not None) else 0.0
                    heading_scores[cid] = max(heading_scores.get(cid, 0.0), s)
                    if cid not in obj_map:
                        obj_map[cid] = o

            all_cids = set(list(text_scores.keys()) + list(heading_scores.keys()))
            max_t = max(text_scores.values(), default=1.0) or 1.0
            max_h = max(heading_scores.values(), default=1.0) or 1.0

            rows = []
            for cid in all_cids:
                t        = text_scores.get(cid, 0.0) / max_t
                h        = (heading_scores.get(cid, 0.0) / max_h) if include_ancestral else 0.0
                combined = (W_TEXT * t + W_HEADING * h) if include_ancestral else t
                obj      = obj_map.get(cid)
                doc      = (obj.properties.get("doc_source", "") if obj else "") or chunk_doc_map.get(cid, "UNKNOWN")
                text_val = (obj.properties.get("text", "") if obj else "") or ""
                rows.append({
                    "chunk_id":           cid,
                    "doc_source":         doc,
                    "text_score_norm":    round(t, 6),
                    "heading_score_norm": round(h, 6),
                    "combined_score":     round(combined, 6),
                    "text_snippet":       text_val[:300],
                })

            rows.sort(key=lambda r: r["combined_score"], reverse=True)
            return [{"rank": rank, **row} for rank, row in enumerate(rows[:TOP_N], 1)]

        # ========================
        # RANKING HELPERS
        # ========================
        def first_relevant_rank(results: list[dict], ids: set[str]) -> int | None:
            for r in results:
                if r["chunk_id"] in ids:
                    return r["rank"]
            return None

        def _mrr(results: list[dict], ids: set[str]) -> float:
            for r in results:
                if r["chunk_id"] in ids:
                    return round(1.0 / r["rank"], 6)
            return 0.0

        # ========================
        # PER-DOCUMENT EVALUATION
        # ========================
        all_results: list[dict] = []

        # Flatten doc_meta × queries into (meta, query, query_type) triples
        query_triples: list[tuple[dict, str, str]] = []
        for meta in doc_meta:
            query_triples.append((meta, meta["root_heading"], "title"))
            for qi, eq in enumerate(meta["extra_questions"], 1):
                query_triples.append((meta, eq, f"extra_{qi}"))

        total_q = len(query_triples)
        print(f"\nRunning queries ({len(doc_meta)} docs, {total_q} total queries, TOP_N={TOP_N}) ...")
        print("-" * 70)

        for sno, (meta, query, query_type) in enumerate(query_triples, 1):
            stem       = meta["file_stem"]
            gt_ids     = meta["gt_chunk_ids"]
            gt_count   = meta["gt_count"]

            print(f"[{sno}/{total_q}] {stem}  [{query_type}]")
            print(f"  Query : {query}")

            exec_errors = ""
            with_anc: list[dict] = []
            without_anc: list[dict] = []

            try:
                with_anc    = run_retrieval(query, include_ancestral=True)
                without_anc = run_retrieval(query, include_ancestral=False)
            except Exception as e:
                exec_errors = str(e)
                print(f"  [ERROR] {e}")

            ground_truth_json = json.dumps(sorted(gt_ids))

            wa_ids = {r["chunk_id"] for r in with_anc}
            wo_ids = {r["chunk_id"] for r in without_anc}

            overlap_with    = len(gt_ids & wa_ids)
            overlap_without = len(gt_ids & wo_ids)

            is_doc_retrieved_with    = overlap_with    > 0
            is_doc_retrieved_without = overlap_without > 0

            doc_chunks_in_topN_with    = sum(1 for r in with_anc    if r["chunk_id"] in gt_ids)
            doc_chunks_in_topN_without = sum(1 for r in without_anc if r["chunk_id"] in gt_ids)

            doc_title_in_top1_with    = bool(with_anc    and with_anc[0]["chunk_id"]    in gt_ids)
            doc_title_in_top1_without = bool(without_anc and without_anc[0]["chunk_id"] in gt_ids)

            top1_wa = with_anc[0]["chunk_id"]    if with_anc    else None
            top1_wo = without_anc[0]["chunk_id"] if without_anc else None
            same_chunk_1    = (top1_wa == top1_wo) if (top1_wa and top1_wo) else False
            same_all_chunks = (wa_ids == wo_ids)
            same_chunk_count = len(wa_ids & wo_ids)

            top2_wa = {r["chunk_id"] for r in with_anc[:2]}
            top2_wo = {r["chunk_id"] for r in without_anc[:2]}
            top3_wa = {r["chunk_id"] for r in with_anc[:3]}
            top3_wo = {r["chunk_id"] for r in without_anc[:3]}
            common_chunks_top2 = json.dumps(sorted(top2_wa & top2_wo))
            common_chunks_top3 = json.dumps(sorted(top3_wa & top3_wo))

            frr_with    = first_relevant_rank(with_anc,    gt_ids)
            frr_without = first_relevant_rank(without_anc, gt_ids)

            recall_with    = round(overlap_with    / gt_count, 6) if gt_count else 0.0
            recall_without = round(overlap_without / gt_count, 6) if gt_count else 0.0

            k = TOP_N
            prec_with    = round(doc_chunks_in_topN_with    / k, 6) if k else 0.0
            prec_without = round(doc_chunks_in_topN_without / k, 6) if k else 0.0
            delta_prec   = round(prec_with - prec_without, 6)

            mrr_with    = _mrr(with_anc,    gt_ids)
            mrr_without = _mrr(without_anc, gt_ids)

            relevancy_with    = round(overlap_with    / len(with_anc),    6) if with_anc    else 0.0
            relevancy_without = round(overlap_without / len(without_anc), 6) if without_anc else 0.0
            delta_relevancy   = round(relevancy_with - relevancy_without, 6)

            if delta_prec > 0:
                anc_better = "YES"
            elif delta_prec == 0:
                anc_better = "TIE"
            else:
                anc_better = "NO"

            issue = "ISSUE" if anc_better == "NO" else ""

            print(f"  Prec@{k}: With={prec_with:.0%}  Without={prec_without:.0%}  "
                  f"Delta={delta_prec:+.0%}  → {anc_better}")

            all_results.append({
                "SNo":                    sno,
                "query_type":             query_type,
                "question":               query,
                "document_source":        stem,
                "ground_truth":           ground_truth_json,
                "Ground_truth_chunk_count": gt_count,
                "retrieved_count_with_anc":    len(with_anc),
                "retrieved_count_without_anc": len(without_anc),
                "with_anc":    with_anc,
                "without_anc": without_anc,
                "gt_ids":      gt_ids,
                "is_doc_retrieved_with_anc":    is_doc_retrieved_with,
                "is_doc_retrieved_without_anc": is_doc_retrieved_without,
                "overlap_count_with_anc":    overlap_with,
                "overlap_count_without_anc": overlap_without,
                "doc_title_in_top1_with_anc":    doc_title_in_top1_with,
                "doc_title_in_top1_without_anc": doc_title_in_top1_without,
                "doc_chunks_in_topN_with_anc":    doc_chunks_in_topN_with,
                "doc_chunks_in_topN_without_anc": doc_chunks_in_topN_without,
                "same_chunk_1":      same_chunk_1,
                "same_all_chunks":   same_all_chunks,
                "same_chunk_count":  same_chunk_count,
                "common_chunks_top2": common_chunks_top2,
                "common_chunks_top3": common_chunks_top3,
                "first_relevant_rank_with_anc":    frr_with,
                "first_relevant_rank_without_anc": frr_without,
                "recall_at_n_with_anc":    recall_with,
                "recall_at_n_without_anc": recall_without,
                "precision_at_k_with_anc":    prec_with,
                "precision_at_k_without_anc": prec_without,
                "delta_precision": delta_prec,
                "MRR_with_anc":    mrr_with,
                "MRR_without_anc": mrr_without,
                "relevancy_with_anc":    relevancy_with,
                "relevancy_without_anc": relevancy_without,
                "delta_relevancy":       delta_relevancy,
                "anc_better":  anc_better,
                "executionErrors": exec_errors,
                "Issue":           issue,
            })

        print("-" * 70)
        n = len(all_results)
        avg_prec_with      = sum(r["precision_at_k_with_anc"]    for r in all_results) / n if n else 0
        avg_prec_without   = sum(r["precision_at_k_without_anc"] for r in all_results) / n if n else 0
        avg_delta          = avg_prec_with - avg_prec_without
        avg_recall_with    = sum(r["recall_at_n_with_anc"]    for r in all_results) / n if n else 0
        avg_recall_without = sum(r["recall_at_n_without_anc"] for r in all_results) / n if n else 0
        avg_mrr_with       = sum(r["MRR_with_anc"]    for r in all_results) / n if n else 0
        avg_mrr_without    = sum(r["MRR_without_anc"] for r in all_results) / n if n else 0
        avg_relevancy_with    = sum(r["relevancy_with_anc"]    for r in all_results) / n if n else 0
        avg_relevancy_without = sum(r["relevancy_without_anc"] for r in all_results) / n if n else 0
        avg_delta_relevancy   = avg_relevancy_with - avg_relevancy_without
        count_yes = sum(1 for r in all_results if r["anc_better"] == "YES")
        count_tie = sum(1 for r in all_results if r["anc_better"] == "TIE")
        count_no  = sum(1 for r in all_results if r["anc_better"] == "NO")

        print(f"\nSUMMARY ACROSS {n} QUERIES ({len(doc_meta)} documents):")
        print(f"  Config          : alpha={ALPHA}  w_text={W_TEXT}  w_heading={W_HEADING}  TOP_N={TOP_N}")
        print(f"  Avg Precision@{TOP_N}: With={avg_prec_with:.2%}  Without={avg_prec_without:.2%}  Delta={avg_delta:+.2%}")
        print(f"  Avg Recall@{TOP_N}   : With={avg_recall_with:.2%}  Without={avg_recall_without:.2%}")
        print(f"  Avg MRR         : With={avg_mrr_with:.4f}  Without={avg_mrr_without:.4f}")
        print(f"  Avg Relevancy   : With={avg_relevancy_with:.2%}  Without={avg_relevancy_without:.2%}  Delta={avg_delta_relevancy:+.2%}")
        print(f"  Ancestral wins  : {count_yes} YES / {count_tie} TIE / {count_no} NO")
        print("-" * 70)

        # ========================
        # WRITE EXCEL
        # ========================
        _write_excel(
            all_results=all_results,
            excel_path=excel_path,
            W_TEXT=W_TEXT,
            W_HEADING=W_HEADING,
            TOP_N=TOP_N,
            avg_recall_with=avg_recall_with,
            avg_recall_without=avg_recall_without,
            avg_prec_with=avg_prec_with,
            avg_prec_without=avg_prec_without,
            avg_delta=avg_delta,
            avg_mrr_with=avg_mrr_with,
            avg_mrr_without=avg_mrr_without,
            avg_relevancy_with=avg_relevancy_with,
            avg_relevancy_without=avg_relevancy_without,
            avg_delta_relevancy=avg_delta_relevancy,
            count_yes=count_yes,
            count_tie=count_tie,
            count_no=count_no,
        )

    finally:
        client.close()
        print("Done.")


# ========================
# EXCEL WRITER (separated for clarity)
# ========================
def _write_excel(
    *,
    all_results: list[dict],
    excel_path: Path,
    W_TEXT: float,
    W_HEADING: float,
    TOP_N: int,
    avg_recall_with: float,
    avg_recall_without: float,
    avg_prec_with: float,
    avg_prec_without: float,
    avg_delta: float,
    avg_mrr_with: float,
    avg_mrr_without: float,
    avg_relevancy_with: float,
    avg_relevancy_without: float,
    avg_delta_relevancy: float,
    count_yes: int,
    count_tie: int,
    count_no: int,
) -> None:
    n = len(all_results)
    wb = Workbook()

    # ── Shared Styles ────────────────────────────────────────────────────
    hdr_font     = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill_blu = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_fill_grn = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    hdr_fill_orn = PatternFill(start_color="833C00", end_color="833C00", fill_type="solid")
    sub_font     = Font(bold=True, size=10)
    sub_fill     = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gold_fill  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    bold_font  = Font(bold=True, size=11)
    ctr        = Alignment(horizontal="center", vertical="center")
    wrap_ctr   = Alignment(wrap_text=True, vertical="center", horizontal="center")

    # ====================================================================
    # SHEET 1 — Detailed Results
    # A-F  : Query Info
    # G-N  : With Ancestral (8 cols)
    # O-T  : Without Ancestral (6 cols)
    # ====================================================================
    ws1 = wb.active
    ws1.title = "Detailed Results"

    ws1.merge_cells("A1:F1"); ws1["A1"] = "Query Info"
    ws1.merge_cells("G1:N1"); ws1["G1"] = f"With Ancestral-Heading  (w_text={W_TEXT}, w_heading={W_HEADING})"
    ws1.merge_cells("O1:T1"); ws1["O1"] = "Without Ancestral-Heading  (text score only)"

    for ref, fill in [("A1", hdr_fill_blu), ("G1", hdr_fill_grn), ("O1", hdr_fill_orn)]:
        c = ws1[ref]; c.font = hdr_font; c.fill = fill; c.alignment = ctr; c.border = thin_border

    sub_hdrs = [
        "SNo", "Question  [query_type: title / extra_N]", "Document Source",
        "GT Chunk Count", "Retrieved (W/Anc)", "Retrieved (W/o Anc)",
        "Rank", "Chunk ID", "Doc Source", "Correct?",
        "text_norm", "heading_norm", "combined", "Text Snippet",
        "Rank", "Chunk ID", "Doc Source", "Correct?",
        "text_norm", "Text Snippet",
    ]
    for ci, h in enumerate(sub_hdrs, 1):
        c = ws1.cell(row=2, column=ci, value=h)
        c.font = sub_font; c.fill = sub_fill; c.border = thin_border; c.alignment = ctr

    ws1.row_dimensions[1].height = 22
    ws1.row_dimensions[2].height = 28

    cur_row = 3
    for res in all_results:
        with_anc_rows    = res["with_anc"]
        without_anc_rows = res["without_anc"]
        n_rows   = max(len(with_anc_rows), len(without_anc_rows), 1)
        gt_ids_r = res["gt_ids"]

        ws1.cell(row=cur_row, column=1, value=res["SNo"])
        ws1.cell(row=cur_row, column=2, value=f"[{res['query_type']}] {res['question']}")
        ws1.cell(row=cur_row, column=3, value=res["document_source"])
        ws1.cell(row=cur_row, column=4, value=res["Ground_truth_chunk_count"])
        ws1.cell(row=cur_row, column=5, value=res["retrieved_count_with_anc"])
        ws1.cell(row=cur_row, column=6, value=res["retrieved_count_without_anc"])

        for i in range(n_rows):
            row = cur_row + i

            if i < len(with_anc_rows):
                wa      = with_anc_rows[i]
                correct = wa["chunk_id"] in gt_ids_r
                for ci, v in enumerate([
                    wa["rank"], wa["chunk_id"], wa["doc_source"],
                    "YES" if correct else "no",
                    wa["text_score_norm"], wa["heading_score_norm"], wa["combined_score"],
                    wa["text_snippet"],
                ], 7):
                    ws1.cell(row=row, column=ci, value=v)
                fill = green_fill if correct else red_fill
                for col in range(7, 15):
                    ws1.cell(row=row, column=col).fill = fill

            if i < len(without_anc_rows):
                wo      = without_anc_rows[i]
                correct = wo["chunk_id"] in gt_ids_r
                for ci, v in enumerate([
                    wo["rank"], wo["chunk_id"], wo["doc_source"],
                    "YES" if correct else "no",
                    wo["text_score_norm"],
                    wo["text_snippet"],
                ], 15):
                    ws1.cell(row=row, column=ci, value=v)
                fill = green_fill if correct else red_fill
                for col in range(15, 21):
                    ws1.cell(row=row, column=col).fill = fill

            for col in range(1, 21):
                ws1.cell(row=row, column=col).border = thin_border

        if n_rows > 1:
            for col in range(1, 7):
                ws1.merge_cells(
                    start_row=cur_row, start_column=col,
                    end_row=cur_row + n_rows - 1, end_column=col,
                )
                ws1.cell(row=cur_row, column=col).alignment = wrap_ctr

        cur_row += n_rows

    for col_letter, w in {
        "A": 6, "B": 42, "C": 36, "D": 12, "E": 14, "F": 16,
        "G": 6, "H": 44, "I": 32, "J": 9, "K": 11, "L": 13, "M": 11, "N": 40,
        "O": 6, "P": 44, "Q": 32, "R": 9, "S": 11, "T": 40,
    }.items():
        ws1.column_dimensions[col_letter].width = w

    # ====================================================================
    # SHEET 2 — Summary
    # ====================================================================
    ws2 = wb.create_sheet("Summary")

    sum_col_groups = {
        "Identity":           (1,  ["SNo", "Question  [query_type: title / extra_N]", "Document Source", "GT Chunk Count"]),
        "GT Presence":        (5,  ["Retrieved (W/Anc)", "Retrieved (W/o Anc)",
                                    "is_retrieved_with", "is_retrieved_without",
                                    "overlap_with", "overlap_without"]),
        "Doc Title Match":    (11, ["top1_correct_with", "top1_correct_without",
                                    "topN_correct_with", "topN_correct_without"]),
        "Cross-Mode Overlap": (15, ["same_top1", "same_all_chunks",
                                    "same_chunk_count", "common_top2", "common_top3"]),
        "Ranking Metrics":    (20, ["first_rank_with", "first_rank_without",
                                    "recall@N_with", "recall@N_without"]),
        "Derived Metrics":    (24, [f"Prec@{TOP_N}_with", f"Prec@{TOP_N}_without",
                                    "delta_precision", "MRR_with", "MRR_without",
                                    "relevancy_with", "relevancy_without", "delta_relevancy"]),
        "Result":             (32, ["anc_better", "executionErrors", "Issue"]),
    }

    fill_map = {
        "Identity":           PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "GT Presence":        PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
        "Doc Title Match":    PatternFill(start_color="375623", end_color="375623", fill_type="solid"),
        "Cross-Mode Overlap": PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid"),
        "Ranking Metrics":    PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid"),
        "Derived Metrics":    PatternFill(start_color="843C0C", end_color="843C0C", fill_type="solid"),
        "Result":             PatternFill(start_color="375623", end_color="375623", fill_type="solid"),
    }

    for group, (start_col, cols) in sum_col_groups.items():
        end_col = start_col + len(cols) - 1
        ws2.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        c = ws2.cell(row=1, column=start_col, value=group)
        c.font = hdr_font; c.fill = fill_map[group]; c.alignment = ctr; c.border = thin_border

    all_sum_cols: list[str] = []
    for _g, (_s, cols) in sum_col_groups.items():
        all_sum_cols.extend(cols)

    for ci, h in enumerate(all_sum_cols, 1):
        c = ws2.cell(row=2, column=ci, value=h)
        c.font = sub_font; c.fill = sub_fill; c.border = thin_border; c.alignment = ctr

    ws2.row_dimensions[1].height = 22
    ws2.row_dimensions[2].height = 32

    for ri, res in enumerate(all_results, 3):
        vals = [
            res["SNo"], f"[{res['query_type']}] {res['question']}", res["document_source"], res["Ground_truth_chunk_count"],
            res["retrieved_count_with_anc"], res["retrieved_count_without_anc"],
            "YES" if res["is_doc_retrieved_with_anc"]    else "no",
            "YES" if res["is_doc_retrieved_without_anc"] else "no",
            res["overlap_count_with_anc"], res["overlap_count_without_anc"],
            "YES" if res["doc_title_in_top1_with_anc"]    else "no",
            "YES" if res["doc_title_in_top1_without_anc"] else "no",
            res["doc_chunks_in_topN_with_anc"], res["doc_chunks_in_topN_without_anc"],
            "YES" if res["same_chunk_1"]   else "no",
            "YES" if res["same_all_chunks"] else "no",
            res["same_chunk_count"], res["common_chunks_top2"], res["common_chunks_top3"],
            res["first_relevant_rank_with_anc"]    if res["first_relevant_rank_with_anc"]    is not None else "—",
            res["first_relevant_rank_without_anc"] if res["first_relevant_rank_without_anc"] is not None else "—",
            res["recall_at_n_with_anc"], res["recall_at_n_without_anc"],
            res["precision_at_k_with_anc"], res["precision_at_k_without_anc"],
            res["delta_precision"], res["MRR_with_anc"], res["MRR_without_anc"],
            res["relevancy_with_anc"], res["relevancy_without_anc"], res["delta_relevancy"],
            res["anc_better"], res["executionErrors"], res["Issue"],
        ]

        for ci, v in enumerate(vals, 1):
            c = ws2.cell(row=ri, column=ci, value=v)
            c.border = thin_border
            c.alignment = Alignment(
                wrap_text=True, vertical="center",
                horizontal="left" if ci == 2 else "center",
            )

        row_fill = (green_fill if res["anc_better"] == "YES"
                    else (red_fill if res["anc_better"] == "NO" else gold_fill))
        for col in range(11, 33):
            ws2.cell(row=ri, column=col).fill = row_fill

    avg_row = len(all_results) + 4
    footer_vals = {
        1:  "AVERAGE",
        2:  f"{n} documents | TOP_N = {TOP_N} | w_text = {W_TEXT} | w_heading = {W_HEADING}",
        4:  f"{sum(r['Ground_truth_chunk_count'] for r in all_results) / n:.1f}" if n else "",
        5:  f"{sum(r['retrieved_count_with_anc'] for r in all_results) / n:.1f}" if n else "",
        6:  f"{sum(r['retrieved_count_without_anc'] for r in all_results) / n:.1f}" if n else "",
        9:  f"{sum(r['overlap_count_with_anc'] for r in all_results) / n:.2f}" if n else "",
        10: f"{sum(r['overlap_count_without_anc'] for r in all_results) / n:.2f}" if n else "",
        13: f"{sum(r['doc_chunks_in_topN_with_anc'] for r in all_results) / n:.2f}" if n else "",
        14: f"{sum(r['doc_chunks_in_topN_without_anc'] for r in all_results) / n:.2f}" if n else "",
        17: f"{sum(r['same_chunk_count'] for r in all_results) / n:.2f}" if n else "",
        22: f"{avg_recall_with:.2%}",
        23: f"{avg_recall_without:.2%}",
        24: f"{avg_prec_with:.2%}",
        25: f"{avg_prec_without:.2%}",
        26: f"{avg_delta:+.2%}",
        27: f"{avg_mrr_with:.4f}",
        28: f"{avg_mrr_without:.4f}",
        29: f"{avg_relevancy_with:.2%}",
        30: f"{avg_relevancy_without:.2%}",
        31: f"{avg_delta_relevancy:+.2%}",
        32: f"YES:{count_yes} TIE:{count_tie} NO:{count_no}",
    }
    for col, val in footer_vals.items():
        c = ws2.cell(row=avg_row, column=col, value=val)
        c.font = bold_font; c.border = thin_border
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.fill = green_fill if avg_delta > 0 else (red_fill if avg_delta < 0 else gold_fill)

    sum_col_widths = [
        5, 42, 36, 10,
        13, 14, 13, 14, 11, 12,
        13, 14, 13, 14,
        11, 12, 14, 22, 22,
        13, 14, 12, 12,
        13, 13, 14, 12, 12, 13, 13, 14,
        10, 26, 10,
    ]
    for ci, w in enumerate(sum_col_widths, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    for row_idx in range(3, avg_row + 1):
        ws2.row_dimensions[row_idx].height = 32

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(excel_path))
    print(f"\nExcel saved to: {excel_path}")
    print(f"  Sheet 1 'Detailed Results': {cur_row - 3} result rows across {len(all_results)} docs")
    print(f"  Sheet 2 'Summary': {len(all_results)} document rows + average footer")

