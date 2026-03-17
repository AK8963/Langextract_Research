"""
make_results_excel.py
─────────────────────
Generates results/rag_results.xlsx from the captured run data.
Run once:  python make_results_excel.py
"""
from openpyxl import Workbook
from openpyxl.styles import ( PatternFill, Font, Alignment, Border, Side )
from openpyxl.utils import get_column_letter
from pathlib import Path
from collections.abc import Mapping
import json

C_HEADER_DARK   = "1F3864"   # dark navy  – section headers
C_HEADER_MID    = "2E75B6"   # mid blue   – column headers
C_ACCENT        = "D6E4F0"   # light blue – alternate rows / sub-headers
C_GREEN_LIGHT   = "E2EFDA"   # light green – success cells
C_YELLOW        = "FFF2CC"   # yellow – timing cells
C_WHITE         = "FFFFFF"
C_BORDER        = "BDD7EE"

thin  = Side(style="thin",   color=C_BORDER)
thick = Side(style="medium", color=C_HEADER_MID)

BORDER_ALL  = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
BORDER_TOP  = Border(left=thin,  right=thin,  top=thick, bottom=thin)

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, white=False, size=11, italic=False) -> Font:
    color = "FFFFFF" if white else "000000"
    return Font(bold=bold, color=color, size=size, italic=italic)

def _center(wrap=False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap=False) -> Alignment:
    return Alignment(horizontal="left", vertical="top", wrap_text=wrap)

def _set_col_widths(ws, widths: Mapping[int, int | float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def _merge_header(ws, row, col_start, col_end, text, bg=C_HEADER_DARK) -> None:
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.fill      = _fill(bg)
    cell.font      = _font(bold=True, white=True, size=12)
    cell.alignment = _center(wrap=False)
    cell.border    = BORDER_TOP

def _col_header(ws, row, col, text, bg=C_HEADER_MID) -> None:
    c = ws.cell(row=row, column=col, value=text)
    c.fill      = _fill(bg)
    c.font      = _font(bold=True, white=True, size=10)
    c.alignment = _center(wrap=True)
    c.border    = BORDER_ALL

def _data_cell(ws, row, col, value, bg=C_WHITE, bold=False, align="left", wrap=True, number_format=None) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = _fill(bg)
    c.font      = _font(bold=bold, size=10)
    c.alignment = _left(wrap) if align == "left" else _center(wrap)
    c.border    = BORDER_ALL
    if number_format:
        c.number_format = number_format

METRICS = {
    "document_name": "",
    "document_length": 0,
    "chunk_size": 0,
    "total_chunks_processed": 0,
    "raw_headings_extracted": 0,
    "valid_unique_headings": 0,
    "rejected_false_pattern": 0,
    "rejected_body_text": 0,
    "rejected_duplicate": 0,
    "rejected_not_in_text": 0,
    "output_chunks_created": 0,
    "llm_successes": 0,
    "regex_fallbacks_used": 0,
    "llm_calls": 0,
    "llm_input_tokens": 0,
    "llm_output_tokens": 0,
    "doc_retrieval_time_s": 0.0,
    "claims_extraction_time_s": 0.0,
    "llm_answering_time_s": 0.0,
    "llm_per_chunk_times_s": "",
    "llm_total_processing_time_s": 0.0,
    "validation_rate_pct": "0.0%",
}

QUERIES = []

def generate_excel_report(run_metrics: dict, output_json_path: str) -> str:
    """
    Generate Excel report from run metrics.
    Args:
        run_metrics: Dictionary with all run metrics from main.py
        output_json_path: Path to the output JSON (used to derive Excel path)
    Returns:
        Path to generated Excel file
    """
    METRICS['document_name'] = run_metrics.get('document_name', 'unknown.pdf')
    METRICS['document_length'] = run_metrics.get('doc_length', 0)
    METRICS['chunk_size'] = run_metrics.get('chunk_size', 0)
    METRICS['total_chunks_processed'] = run_metrics.get('total_chunks', 0)
    METRICS['raw_headings_extracted'] = run_metrics.get('raw_headings', 0)
    METRICS['valid_unique_headings'] = run_metrics.get('valid_headings', 0)
    METRICS['rejected_false_pattern'] = run_metrics.get('rejected_false', 0)
    METRICS['rejected_body_text'] = run_metrics.get('rejected_body', 0)
    METRICS['rejected_duplicate'] = run_metrics.get('rejected_duplicate', 0)
    METRICS['rejected_not_in_text'] = run_metrics.get('rejected_not_found', 0)
    METRICS['output_chunks_created'] = run_metrics.get('output_chunks', 0)
    METRICS['llm_successes'] = run_metrics.get('llm_success', 0)
    METRICS['regex_fallbacks_used'] = run_metrics.get('regex_fallbacks', 0)
    METRICS['llm_calls'] = run_metrics.get('llm_calls', 0)
    METRICS['llm_input_tokens'] = run_metrics.get('llm_input_tokens', 0)
    METRICS['llm_output_tokens'] = run_metrics.get('llm_output_tokens', 0)
    METRICS['doc_retrieval_time_s'] = round(run_metrics.get('doc_retrieval_time', 0), 2)
    METRICS['claims_extraction_time_s'] = round(run_metrics.get('claims_extraction_time', 0), 2)
    METRICS['llm_answering_time_s'] = round(run_metrics.get('llm_total_time', 0), 2)
    METRICS['llm_total_processing_time_s'] = round(run_metrics.get('llm_total_time', 0), 2)
    METRICS['validation_rate_pct'] = f"{run_metrics.get('validation_rate', 0):.1f}%"

    chunk_times = run_metrics.get('llm_chunk_times', [])
    if chunk_times:
        avg_time = sum(c['time_s'] for c in chunk_times) / len(chunk_times)
        METRICS['llm_per_chunk_times_s'] = f"Avg: {avg_time:.2f}s ({len(chunk_times)} chunks)"
    else:
        METRICS['llm_per_chunk_times_s'] = "N/A"

    wb = Workbook()
    if wb.active is not None:
        wb.remove(wb.active)
    ws1 = wb.create_sheet("Overview")
    ws1.sheet_properties.tabColor = "1F3864"
    ws1.freeze_panes = "A4"
    r = 1
    ws1.merge_cells(f"A{r}:L{r}")
    c = ws1.cell(row=r, column=1, value="PDF Ancestral Heading Extraction Pipeline — Run Metrics")
    c.fill = _fill(C_HEADER_DARK); c.font = _font(bold=True, white=True, size=14)
    c.alignment = _center(); c.border = BORDER_TOP
    ws1.row_dimensions[r].height = 26
    r += 1
    ws1.merge_cells(f"A{r}:L{r}")
    model_name = run_metrics.get('model', 'gemma2:9b')
    c = ws1.cell(row=r, column=1, value=f"Document: {METRICS['document_name']}  |  Model: {model_name}")
    c.fill = _fill(C_ACCENT); c.font = _font(italic=True, size=10)
    c.alignment = _center(); c.border = BORDER_ALL
    r += 1

    # ── RUN METRICS ─────────────────────────────────────────────────────────────
    _merge_header(ws1, r, 1, 6, "RUN METRICS", C_HEADER_DARK)
    _merge_header(ws1, r, 7, 12, "LLM & Timing", C_HEADER_DARK)
    r += 1

    metrics_left = [
    ("Document Name", METRICS["document_name"], C_WHITE),
    ("Document Length (chars)", METRICS["document_length"], C_WHITE),
    ("Chunk Size", METRICS["chunk_size"], C_WHITE),
    ("Total Chunks Processed", METRICS["total_chunks_processed"], C_WHITE),
    ("Raw Headings Extracted", METRICS["raw_headings_extracted"], C_WHITE),
    ("Valid Unique Headings", METRICS["valid_unique_headings"], C_GREEN_LIGHT),
    ("Rejected (False Pattern)", METRICS["rejected_false_pattern"], C_WHITE),
    ("Rejected (Body Text)", METRICS["rejected_body_text"], C_WHITE),
    ("Rejected (Duplicate)", METRICS["rejected_duplicate"], C_WHITE),
    ("Rejected (Not in Text)", METRICS["rejected_not_in_text"], C_WHITE),
    ("Output Chunks Created", METRICS["output_chunks_created"], C_WHITE),
    ("LLM Successes", METRICS["llm_successes"], C_WHITE),
    ("Regex Fallbacks Used", METRICS["regex_fallbacks_used"], C_WHITE),
    ("Validation Rate", METRICS["validation_rate_pct"], C_GREEN_LIGHT),
    ]

    metrics_right = [
    ("LLM Calls", METRICS["llm_calls"], C_WHITE),
    ("LLM Input Tokens", METRICS["llm_input_tokens"], C_WHITE),
    ("LLM Output Tokens", METRICS["llm_output_tokens"], C_WHITE),
    ("Doc Retrieval Time (s)", METRICS["doc_retrieval_time_s"], C_YELLOW),
    ("Claims Extraction Time (s)", METRICS["claims_extraction_time_s"], C_YELLOW),
    ("LLM Answering Time (s)", METRICS["llm_answering_time_s"], C_YELLOW),
    ("LLM Total Processing Time (s)", METRICS["llm_total_processing_time_s"], C_YELLOW),
    ("LLM Per-Chunk Times", METRICS["llm_per_chunk_times_s"], C_ACCENT),
    ]

    max_rows = max(len(metrics_left), len(metrics_right))
    for i in range(max_rows):
        left = metrics_left[i] if i < len(metrics_left) else ("", "", C_WHITE)
        right = metrics_right[i] if i < len(metrics_right) else ("", "", C_WHITE)

        _data_cell(ws1, r, 1, left[0],  C_ACCENT, bold=True, align="left")
        _data_cell(ws1, r, 2, left[1],  left[2],  bold=False, align="center")
        ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        _data_cell(ws1, r, 7, right[0], C_ACCENT, bold=True, align="left")
        _data_cell(ws1, r, 8, right[1], right[2], bold=False, align="center")
        ws1.merge_cells(start_row=r, start_column=8, end_row=r, end_column=12)
        ws1.row_dimensions[r].height = 18
        r += 1

    r += 1
    # ── PER-CHUNK LLM METRICS ────────────────────────────────────────────────────
    _merge_header(ws1, r, 1, 12, "PER-CHUNK LLM PROCESSING METRICS", C_HEADER_DARK)
    r += 1

    chunk_headers = ["Chunk #", "Chunk Size (chars)", "Processing Time (s)", "Input Tokens", "Output Tokens", "Headings Found", "Status"]
    for ci, h in enumerate(chunk_headers, 1):
        _col_header(ws1, r, ci, h)
    ws1.row_dimensions[r].height = 20
    r += 1

    # Get chunk times from metrics
    chunk_times = run_metrics.get('llm_chunk_times', [])
    
    if chunk_times:
        for chunk_data in chunk_times:
            _data_cell(ws1, r, 1, chunk_data.get('chunk_index', ''), C_WHITE, align="center")
            _data_cell(ws1, r, 2, chunk_data.get('chunk_size', 0), C_WHITE, align="center")
            _data_cell(ws1, r, 3, chunk_data.get('time_s', 0), C_YELLOW, align="center", bold=True)
            _data_cell(ws1, r, 4, chunk_data.get('input_tokens', 0), C_WHITE, align="center")
            _data_cell(ws1, r, 5, chunk_data.get('output_tokens', 0), C_WHITE, align="center")
            _data_cell(ws1, r, 6, chunk_data.get('headings_found', 0), C_GREEN_LIGHT, align="center")
            _data_cell(ws1, r, 7, "Success", C_GREEN_LIGHT, align="center")
            ws1.row_dimensions[r].height = 18
            r += 1

        r += 1
        _data_cell(ws1, r, 1, "TOTAL", C_HEADER_MID, bold=True, align="center")
        total_chunk_size = sum(c.get('chunk_size', 0) for c in chunk_times)
        total_time = sum(c.get('time_s', 0) for c in chunk_times)
        total_input = sum(c.get('input_tokens', 0) for c in chunk_times)
        total_output = sum(c.get('output_tokens', 0) for c in chunk_times)
        total_headings = sum(c.get('headings_found', 0) for c in chunk_times)
        _data_cell(ws1, r, 2, total_chunk_size, C_ACCENT, bold=True, align="center")
        _data_cell(ws1, r, 3, round(total_time, 2), C_YELLOW, bold=True, align="center")
        _data_cell(ws1, r, 4, total_input, C_ACCENT, bold=True, align="center")
        _data_cell(ws1, r, 5, total_output, C_ACCENT, bold=True, align="center")
        _data_cell(ws1, r, 6, total_headings, C_GREEN_LIGHT, bold=True, align="center")
        _data_cell(ws1, r, 7, f"{len(chunk_times)} chunks", C_ACCENT, bold=True, align="center")
        ws1.row_dimensions[r].height = 20

    # Set column widths for Overview sheet
    ws1.column_dimensions[get_column_letter(1)].width  = 4
    ws1.column_dimensions[get_column_letter(2)].width  = 46


    # Save to output/metrics_results/ directory
    metrics_dir = Path("output") / "metrics_results"
    metrics_dir.mkdir(parents=True, exist_ok=True)
    
    excel_filename = Path(output_json_path).stem + "_metrics.xlsx"
    out_file = metrics_dir / excel_filename
    wb.save(out_file)
    return str(out_file)


if __name__ == "__main__":
    out = Path("results")
    out.mkdir(exist_ok=True)
    out_file = out / "rag_results.xlsx"
    print(f"Run from main.py for actual metrics integration")
